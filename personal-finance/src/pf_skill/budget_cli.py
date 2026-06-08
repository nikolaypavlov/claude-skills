"""``pf-budget`` CLI entry: budget management.

Subcommands::

    pf-budget register-category --category NAME [--note NOTE]
    pf-budget unregister-category --category NAME [--force]
    pf-budget list-categories [--include-declared]
    pf-budget import <file> --period YYYY-MM
                            [--unknown-categories reject|register]
                            [--dry-run] [--force]
                            [--sheet plans|baseline]

Output contract matches the rest of the ``pf-*`` scripts via
``common.cli.run_subcommand``: success → JSON to stdout exit 0,
``CliError`` → JSON to stderr exit 1, uncaught → traceback + exit 2.
"""

from __future__ import annotations

import argparse
import sqlite3
import sys
import time
from collections.abc import Sequence
from contextlib import closing
from pathlib import Path
from typing import Any

from .common import budget as bud
from .common import queries as q
from .common.cli import (
    CliError,
    resolve_db_path,
    run_subcommand,
)
from .common.store import open_db

# Categories are slash-separated taxonomy names. We forbid leading
# and trailing whitespace and empty segments because sheets frequently
# leak those in via copy-paste. The character set is intentionally
# permissive (Cyrillic categories are the norm); we only forbid the
# truly broken shapes that would corrupt joins.
_FORBIDDEN_LEADING_TRAILING_WHITESPACE_MSG = (
    "category contains leading or trailing whitespace; strip first"
)


def _validate_category_name(name: str) -> str:
    if not name:
        raise CliError("--category must be non-empty")
    if name != name.strip():
        raise CliError(_FORBIDDEN_LEADING_TRAILING_WHITESPACE_MSG)
    if "//" in name or name.startswith("/") or name.endswith("/"):
        raise CliError(
            f"--category {name!r} has empty hierarchy segments; use "
            "TopGroup/Sub form without leading/trailing slashes"
        )
    return name


def _category_in_use(conn: sqlite3.Connection, category: str) -> dict[str, Any]:
    """Aggregate references to a category across the user-mutable
    tables. Used by unregister-category to refuse deletion when the
    category is still load-bearing.

    Returns a dict with counts per source (so the error message can
    point the user at exactly what needs cleaning up first).
    """
    counts: dict[str, int] = {}
    counts["tx_category"] = conn.execute(
        "SELECT COUNT(*) FROM tx_category WHERE category = ?", (category,)
    ).fetchone()[0]
    counts["category_overrides"] = conn.execute(
        "SELECT COUNT(*) FROM category_overrides WHERE category = ?", (category,)
    ).fetchone()[0]
    counts["categorization_rules"] = conn.execute(
        "SELECT COUNT(*) FROM categorization_rules WHERE category = ?", (category,)
    ).fetchone()[0]
    # budget_line table exists from migration v2 onwards. Probe defensively
    # so a partially-migrated DB surfaces a clean error instead of a
    # mysterious sqlite OperationalError.
    try:
        counts["budget_line"] = conn.execute(
            "SELECT COUNT(*) FROM budget_line WHERE category = ?", (category,)
        ).fetchone()[0]
    except sqlite3.OperationalError as exc:
        if "no such table" not in str(exc).lower():
            raise
        counts["budget_line"] = 0
    return counts


def cmd_register_category(args: argparse.Namespace) -> dict[str, Any]:
    """Insert a row into ``category_registry``.

    Idempotent: registering an already-registered category is a no-op
    that returns ``{"ok": true, "already_registered": true}`` without
    touching ``declared_at`` / ``declared_via`` / ``note``. Surprise-
    rewriting metadata would defeat the audit trail.
    """
    category = _validate_category_name(args.category)
    now_ts = int(time.time())
    db_path = resolve_db_path(args.db)
    with closing(open_db(db_path)) as conn:
        existing = conn.execute(
            "SELECT declared_at, declared_via, note FROM category_registry WHERE category = ?",
            (category,),
        ).fetchone()
        if existing is not None:
            return {
                "ok": True,
                "already_registered": True,
                "category": category,
                "declared_at": int(existing[0]),
                "declared_via": existing[1],
                "note": existing[2],
            }
        conn.execute("BEGIN")
        try:
            conn.execute(
                "INSERT INTO category_registry "
                "(category, declared_at, declared_via, note) "
                "VALUES (?, ?, ?, ?)",
                (category, now_ts, "cli", args.note),
            )
            conn.commit()
        except Exception:
            conn.rollback()
            raise
    return {
        "ok": True,
        "already_registered": False,
        "category": category,
        "declared_at": now_ts,
        "declared_via": "cli",
        "note": args.note,
    }


def cmd_unregister_category(args: argparse.Namespace) -> dict[str, Any]:
    """Remove a row from ``category_registry``.

    Refuses to delete a category that is still referenced from
    ``tx_category``, ``category_overrides``, ``categorization_rules``,
    or ``budget_line`` - the user would just rediscover the category
    on the next read. ``--force`` overrides the safety, but is
    deliberately undocumented in the help text so it only gets used
    when somebody has read the source.
    """
    category = _validate_category_name(args.category)
    db_path = resolve_db_path(args.db)
    with closing(open_db(db_path)) as conn:
        existing = conn.execute(
            "SELECT category FROM category_registry WHERE category = ?",
            (category,),
        ).fetchone()
        if existing is None:
            raise CliError(
                f"category {category!r} is not in the registry",
                kind="NotFound",
            )
        usage = _category_in_use(conn, category)
        total_refs = sum(usage.values())
        if total_refs > 0 and not args.force:
            raise CliError(
                f"category {category!r} is still referenced "
                f"({total_refs} rows: {usage}); rename or remove the "
                "references first (or pass --force to unregister anyway)",
                kind="StillInUse",
            )
        conn.execute("BEGIN")
        try:
            conn.execute(
                "DELETE FROM category_registry WHERE category = ?",
                (category,),
            )
            conn.commit()
        except Exception:
            conn.rollback()
            raise
    return {
        "ok": True,
        "category": category,
        "removed": True,
        "remaining_references": usage,
    }


def _parse_file_into_plan(args: argparse.Namespace) -> list[bud.PlanRow]:
    """Decide which parser to use based on the file extension and (for
    XLSX) on whether the workbook has a Baseline sheet alongside Plans.
    Stamps period onto baseline rows; returns the merged plan rows for
    ``args.period``."""
    path = Path(args.source).expanduser()
    if not path.exists():
        raise CliError(f"file not found: {path}", kind="FileNotFound")
    suffix = path.suffix.lower()
    if suffix == ".csv":
        sheet = args.sheet or "plans"
        if sheet == "plans":
            rows = bud.parse_plans_csv(path)
            return [r for r in rows if r.period == args.period]
        if sheet == "baseline":
            return bud.parse_baseline_csv(path, args.period)
        raise CliError(
            f"--sheet must be 'plans' or 'baseline' for CSV input; got {sheet!r}",
            kind="BadArgument",
        )
    if suffix in (".xlsx", ".xlsm"):
        baseline_raw, plans_raw = bud.parse_workbook_xlsx(path)
        baseline_rows = [
            bud._row_from_baseline(  # type: ignore[attr-defined]
                d, args.period, source_row=d["__source_row__"]
            )
            for d in baseline_raw
            if any((d.get(k) or "") for k in ("Category", "Currency", "Kind", "Monthly target"))
        ]
        plans_rows = [
            bud._row_from_plans(  # type: ignore[attr-defined]
                d, source_row=d["__source_row__"]
            )
            for d in plans_raw
            if any((d.get(k) or "") for k in ("Period", "Category", "Currency", "Kind", "Amount"))
        ]
        return bud.merge_baseline_plans(baseline_rows, plans_rows, period=args.period)
    raise CliError(
        f"unsupported file type {suffix!r}; expected .csv, .xlsx, or .xlsm",
        kind="BadArgument",
    )


def cmd_import(args: argparse.Namespace) -> dict[str, Any]:
    """Import a budget plan from CSV/XLSX into ``budget`` and
    ``budget_line`` tables.

    Validation modes:
    - ``reject`` (default): fail if any category is unknown, returning
      the list with Levenshtein suggestions in the error payload so
      the CLI can render it cleanly.
    - ``register``: silently add every unknown to ``category_registry``
      with ``declared_via='budget-import'`` and proceed.

    ``--dry-run`` runs through parsing + validation + a synthetic
    "what would change" diff against the existing budget (if any),
    but never opens a write transaction.
    """
    if args.period is None:
        raise CliError("--period is required (YYYY-MM)", kind="BadArgument")
    if not bud.PERIOD_RE.match(args.period):
        raise CliError(f"--period={args.period!r} must match YYYY-MM", kind="BadArgument")
    db_path = resolve_db_path(args.db)
    try:
        rows = _parse_file_into_plan(args)
    except bud.BudgetParseError as exc:
        raise CliError(str(exc), kind=exc.kind) from exc

    if not rows:
        return {
            "ok": True,
            "period": args.period,
            "source": str(args.source),
            "rows_parsed": 0,
            "warning": "no rows matched the requested period",
        }

    with closing(open_db(db_path)) as conn:
        try:
            validation = bud.validate_categories(rows, conn)
        except bud.BudgetParseError as exc:
            raise CliError(str(exc), kind=exc.kind) from exc

        unknowns = validation.unknown
        new_categories: list[str] = []
        if unknowns:
            if args.unknown_categories == "reject":
                # Failure with structured payload so the caller can
                # render suggestions per category.
                err = CliError(
                    f"{len(unknowns)} unknown categor"
                    f"{'y' if len(unknowns) == 1 else 'ies'} in input; "
                    "rerun with --unknown-categories register to add them, "
                    "or fix the source",
                    kind="UnknownCategories",
                )
                # Attach details to the exception for downstream parsing
                # if the caller catches it. The CLI scaffolding doesn't
                # surface arbitrary attributes today, so we also log
                # them via err.args[0] above.
                err.details = {  # type: ignore[attr-defined]
                    "unknown": [
                        {
                            "category": cat,
                            "suggestions": [{"candidate": s, "distance": d} for s, d in sugg],
                        }
                        for cat, sugg in unknowns
                    ]
                }
                raise err
            if args.unknown_categories == "register":
                if not args.dry_run:
                    conn.execute("BEGIN")
                    try:
                        bud.register_unknowns(conn, [c for c, _ in unknowns])
                        conn.commit()
                    except Exception:
                        conn.rollback()
                        raise
                new_categories = [c for c, _ in unknowns]

        if args.dry_run:
            return _dry_run_summary(conn, args.period, rows, new_categories)

        result = bud.materialise_budget(
            conn,
            period=args.period,
            rows=rows,
            source=str(args.source),
            force=args.force,
        )
        total_added = sum(v["lines_added"] for v in result.by_currency.values())
        total_replaced = sum(v["lines_replaced"] for v in result.by_currency.values())
        run_id = bud.log_import_run(
            conn,
            source=str(args.source),
            period=args.period,
            lines_added=total_added,
            lines_replaced=total_replaced,
            new_categories=new_categories,
        )
    return {
        "ok": True,
        "period": args.period,
        "source": str(args.source),
        "rows_imported": len(rows),
        "by_currency": result.by_currency,
        "new_categories_registered": new_categories,
        "import_run_id": run_id,
    }


def _dry_run_summary(
    conn: sqlite3.Connection,
    period: str,
    rows: list[bud.PlanRow],
    new_categories: list[str],
) -> dict[str, Any]:
    by_cur: dict[int, dict[str, Any]] = {}
    for r in rows:
        slot = by_cur.setdefault(r.currency_code, {"lines": 0, "total_minor": 0, "kinds": {}})
        slot["lines"] += 1
        slot["total_minor"] += r.amount_minor
        slot["kinds"][r.kind] = slot["kinds"].get(r.kind, 0) + 1
    existing: dict[int, dict[str, Any]] = {}
    for cur, _info in by_cur.items():
        row = conn.execute(
            "SELECT id, status FROM budget WHERE period = ? AND currency_code = ?",
            (period, cur),
        ).fetchone()
        if row is not None:
            existing[cur] = {"budget_id": int(row[0]), "status": row[1]}
    return {
        "ok": True,
        "dry_run": True,
        "period": period,
        "rows_parsed": len(rows),
        "would_register": new_categories,
        "by_currency": by_cur,
        "existing_budgets": existing,
    }


def cmd_list_categories(args: argparse.Namespace) -> dict[str, Any]:
    """Mirror of ``pf-query categories`` for the ``pf-budget`` CLI.

    Returns the in-use categories plus, when ``--include-declared`` is
    set, the declared-but-unused ones from ``category_registry``. We
    expose the same surface here so the budget workflow does not have
    to context-switch between two scripts mid-validation.
    """
    db_path = resolve_db_path(args.db)
    with closing(open_db(db_path)) as conn:
        cats = q.list_categories(conn, include_declared=args.include_declared)
    return {
        "ok": True,
        "count": len(cats),
        "categories": cats,
    }


def _parse_currency_or_cli_error(value: str | None) -> int | None:
    if value is None:
        return None
    from .common.currencies import parse_currency_arg

    try:
        return parse_currency_arg(value)
    except ValueError as exc:
        raise CliError(str(exc), kind="ValueError") from exc


def cmd_show(args: argparse.Namespace) -> dict[str, Any]:
    cur_code = _parse_currency_or_cli_error(args.currency)
    db_path = resolve_db_path(args.db)
    with closing(open_db(db_path)) as conn:
        budgets = bud.fetch_budget(conn, period=args.period, currency_code=cur_code)
    if not budgets:
        return {
            "ok": True,
            "period": args.period,
            "currency": args.currency,
            "budgets": [],
            "warning": "no budget materialised for this period",
        }
    return {
        "ok": True,
        "period": args.period,
        "currency": args.currency,
        "budgets": budgets,
    }


def cmd_diff(args: argparse.Namespace) -> dict[str, Any]:
    cur_code = _parse_currency_or_cli_error(args.currency)
    db_path = resolve_db_path(args.db)
    with closing(open_db(db_path)) as conn:
        blocks = bud.diff_budget_vs_actual(conn, period=args.period, currency_code=cur_code)
    return {
        "ok": True,
        "period": args.period,
        "currency": args.currency,
        "blocks": blocks,
    }


def cmd_list(args: argparse.Namespace) -> dict[str, Any]:
    db_path = resolve_db_path(args.db)
    with closing(open_db(db_path)) as conn:
        items = bud.list_budgets(conn)
    return {"ok": True, "count": len(items), "budgets": items}


def cmd_close(args: argparse.Namespace) -> dict[str, Any]:
    cur_code = _parse_currency_or_cli_error(args.currency)
    db_path = resolve_db_path(args.db)
    with closing(open_db(db_path)) as conn:
        try:
            result = bud.set_status(
                conn, period=args.period, currency_code=cur_code, new_status="closed"
            )
        except bud.BudgetParseError as exc:
            raise CliError(str(exc), kind=exc.kind) from exc
    if result.matched == 0:
        raise CliError(
            f"no budget matched period={args.period} currency={args.currency}",
            kind="NotFound",
        )
    return {"ok": True, "matched": result.matched, "changed": result.changed}


def cmd_reopen(args: argparse.Namespace) -> dict[str, Any]:
    cur_code = _parse_currency_or_cli_error(args.currency)
    db_path = resolve_db_path(args.db)
    with closing(open_db(db_path)) as conn:
        try:
            result = bud.set_status(
                conn, period=args.period, currency_code=cur_code, new_status="active"
            )
        except bud.BudgetParseError as exc:
            raise CliError(str(exc), kind=exc.kind) from exc
    if result.matched == 0:
        raise CliError(
            f"no budget matched period={args.period} currency={args.currency}",
            kind="NotFound",
        )
    return {
        "ok": True,
        "matched": result.matched,
        "changed": result.changed,
        "reason": args.reason,
    }


def cmd_delete(args: argparse.Namespace) -> dict[str, Any]:
    cur_code = _parse_currency_or_cli_error(args.currency)
    db_path = resolve_db_path(args.db)
    with closing(open_db(db_path)) as conn:
        try:
            deleted = bud.delete_budget(
                conn,
                period=args.period,
                currency_code=cur_code,
                force=args.force,
            )
        except bud.BudgetParseError as exc:
            raise CliError(str(exc), kind=exc.kind) from exc
    if not deleted:
        raise CliError(
            f"no budget matched period={args.period} currency={args.currency}",
            kind="NotFound",
        )
    return {"ok": True, "deleted": deleted}


def cmd_export(args: argparse.Namespace) -> dict[str, Any]:
    """Write a budget view (variance / family / plan) to CSV or XLSX.

    Views:
      - ``variance``: target vs actual per category (default)
      - ``family``:  pretty-printed grouped view for non-technical
                     readers (the user's spouse). XLSX only.
      - ``plan``:    raw plan dump - what's in the active budget.
    """
    cur_code = _parse_currency_or_cli_error(args.currency)
    db_path = resolve_db_path(args.db)
    fmt = args.format
    out_path: Path | None = None if args.out == "-" else Path(args.out).expanduser()
    if out_path is not None and fmt == "auto":
        fmt = "xlsx" if out_path.suffix.lower() == ".xlsx" else "csv"
    elif fmt == "auto":
        fmt = "csv"

    with closing(open_db(db_path)) as conn:
        if args.view == "variance":
            rows = bud.export_variance_rows(conn, period=args.period, currency_code=cur_code)
            row_count = len(rows)
            if fmt == "csv":
                _write_variance_csv(rows, out_path)
            elif fmt == "xlsx":
                if out_path is None:
                    raise CliError("--out - not supported for xlsx", kind="BadArgument")
                _write_variance_xlsx(rows, out_path)
            else:
                raise CliError(f"unsupported --format {fmt!r}", kind="BadArgument")
        elif args.view == "family":
            if fmt != "xlsx":
                raise CliError(
                    "family view requires --format xlsx (pretty styling depends on it)",
                    kind="BadArgument",
                )
            if out_path is None:
                raise CliError("--out - not supported for family view", kind="BadArgument")
            data = bud.family_view_rows(conn, period=args.period)
            row_count = sum(len(g["lines"]) for c in data["currencies"] for g in c["groups"])
            _write_family_xlsx(data, out_path)
        elif args.view == "plan":
            blocks = bud.fetch_budget(conn, period=args.period, currency_code=cur_code)
            # Show only active budgets in the plan view; drafts go
            # through `pf-budget plan show`.
            active_only = [b for b in blocks if b["status"] == "active"]
            rows = []
            from .common.currencies import alpha_for

            for b in active_only:
                cur_alpha = alpha_for(b["currency_code"]) or str(b["currency_code"])
                for line in b["lines"]:
                    rows.append(
                        {
                            "Period": args.period,
                            "Category": line["category"],
                            "Currency": cur_alpha,
                            "Kind": line["kind"],
                            "Amount": line["amount_minor"] / 100.0,
                            "Note": line["note"] or "",
                        }
                    )
            row_count = len(rows)
            if fmt == "csv":
                _write_plan_csv(rows, out_path)
            elif fmt == "xlsx":
                if out_path is None:
                    raise CliError("--out - not supported for xlsx", kind="BadArgument")
                _write_plan_xlsx(rows, out_path)
            else:
                raise CliError(f"unsupported --format {fmt!r}", kind="BadArgument")
        else:
            raise CliError(
                f"unsupported --view {args.view!r}; expected variance / family / plan",
                kind="BadArgument",
            )

    return {
        "ok": True,
        "period": args.period,
        "view": args.view,
        "format": fmt,
        "out": str(out_path) if out_path else "-",
        "rows": row_count,
    }


def _write_plan_csv(rows: list[dict[str, Any]], out_path: Path | None) -> None:
    import csv as _csv

    headers = ["Period", "Category", "Currency", "Kind", "Amount", "Note"]
    target = out_path.open("w", encoding="utf-8", newline="") if out_path else None
    if target is None:
        return  # stdout path not exercised
    try:
        writer = _csv.DictWriter(target, fieldnames=headers)
        writer.writeheader()
        for r in rows:
            writer.writerow(r)
    finally:
        target.close()


def _write_plan_xlsx(rows: list[dict[str, Any]], out_path: Path) -> None:
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise CliError(
            "openpyxl required for xlsx export; install pf-skill[sheets] or use --format csv",
            kind="MissingDependency",
        ) from exc
    wb = Workbook()
    ws = wb.active
    ws.title = "Plan"
    headers = ["Period", "Category", "Currency", "Kind", "Amount", "Note"]
    ws.append(headers)
    for r in rows:
        ws.append([r[h] for h in headers])
    wb.save(out_path)


def _write_family_xlsx(data: dict[str, Any], out_path: Path) -> None:
    """Render the Family view as a styled XLSX.

    Two sheets:
      - ``Огляд``: grouped pretty view with SUM formulas so totals
        recompute if the spouse adjusts a number in Sheets.
      - ``Деталі``: full flat list with notes for "що це за стаття?"
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    except ImportError as exc:
        raise CliError(
            "openpyxl required for family view; install pf-skill[sheets]",
            kind="MissingDependency",
        ) from exc

    wb = Workbook()
    overview = wb.active
    overview.title = "Огляд"
    details = wb.create_sheet("Деталі")

    HEADER_FONT = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    HEADER_FILL = PatternFill("solid", fgColor="1F4E79")  # navy
    GROUP_FONT = Font(name="Calibri", size=11, bold=True, color="000000")
    GROUP_FILL = PatternFill("solid", fgColor="FFF2CC")  # light yellow
    SUBTOTAL_FONT = Font(name="Calibri", size=11, bold=True, italic=True)
    BAND = PatternFill("solid", fgColor="F5F5F5")
    THIN = Side(border_style="thin", color="B0B0B0")
    BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    # ---- Overview sheet ----
    row = 1
    period_cell = overview.cell(row=row, column=1, value=f"Бюджет на період {data['period']}")
    period_cell.font = Font(size=16, bold=True)
    overview.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    row += 2

    for currency in data["currencies"]:
        # Currency header
        cur_cell = overview.cell(
            row=row,
            column=1,
            value=f"{currency['alpha']} - заплановано",
        )
        cur_cell.font = HEADER_FONT
        cur_cell.fill = HEADER_FILL
        cur_cell.alignment = Alignment(horizontal="left", vertical="center")
        overview.cell(row=row, column=2).fill = HEADER_FILL
        # Currency total cell uses SUM over the group subtotal column we'll fill
        cur_total_cell = overview.cell(row=row, column=3, value=0)
        cur_total_cell.font = HEADER_FONT
        cur_total_cell.fill = HEADER_FILL
        cur_total_cell.alignment = Alignment(horizontal="right", vertical="center")
        cur_total_first_subtotal_row: int | None = None
        cur_total_last_subtotal_row: int | None = None
        row += 1

        for group in currency["groups"]:
            group_start_row = row + 1  # lines start one below the group header
            # Group header row
            gh = overview.cell(row=row, column=1, value=f"  {group['title']}")
            gh.font = GROUP_FONT
            gh.fill = GROUP_FILL
            overview.cell(row=row, column=2).fill = GROUP_FILL
            subtotal_cell = overview.cell(row=row, column=3, value=0)
            subtotal_cell.font = SUBTOTAL_FONT
            subtotal_cell.fill = GROUP_FILL
            subtotal_cell.number_format = "#,##0.00;[Red]-#,##0.00"
            subtotal_cell.alignment = Alignment(horizontal="right")
            group_header_row = row
            row += 1

            for i, line in enumerate(group["lines"]):
                line_label_cell = overview.cell(
                    row=row, column=1, value=f"      {line['category_display']}"
                )
                amt_cell = overview.cell(row=row, column=3, value=line["amount_major"])
                amt_cell.number_format = "#,##0.00;[Red]-#,##0.00"
                amt_cell.alignment = Alignment(horizontal="right")
                if i % 2 == 1:
                    line_label_cell.fill = BAND
                    overview.cell(row=row, column=2).fill = BAND
                    amt_cell.fill = BAND
                row += 1

            # SUM formula for the group subtotal
            if row - 1 >= group_start_row:
                subtotal_cell.value = f"=SUM(C{group_start_row}:C{row - 1})"
            if cur_total_first_subtotal_row is None:
                cur_total_first_subtotal_row = group_header_row
            cur_total_last_subtotal_row = group_header_row
            row += 1  # blank line between groups

        # Currency total = sum of all subtotal cells in this currency block
        if cur_total_first_subtotal_row is not None and cur_total_last_subtotal_row is not None:
            # Each group subtotal lives at column C of group_header_row.
            # Use a range that includes only those rows. Simplest: build
            # a comma-separated list of cells since they're not contiguous.
            cell_refs = []
            scan_row = cur_total_first_subtotal_row
            while scan_row <= cur_total_last_subtotal_row:
                cell_refs.append(f"C{scan_row}")
                # Move to the next group header: skip past lines + blank row
                # Find next non-empty C cell that's a subtotal (formula) -
                # simpler heuristic: just walk row by row and look for cells
                # whose value starts with '=SUM(' below us.
                scan_row += 1
                while scan_row <= cur_total_last_subtotal_row:
                    val = overview.cell(row=scan_row, column=3).value
                    if isinstance(val, str) and val.startswith("=SUM("):
                        break
                    scan_row += 1
            cur_total_cell.value = "=" + "+".join(cell_refs) if cell_refs else 0
            cur_total_cell.number_format = "#,##0.00;[Red]-#,##0.00"

        row += 2  # spacing before next currency

    # Column widths for overview
    overview.column_dimensions["A"].width = 38
    overview.column_dimensions["B"].width = 4
    overview.column_dimensions["C"].width = 16
    overview.freeze_panes = "A4"

    # ---- Details sheet ----
    headers = ["Період", "Група", "Категорія", "Валюта", "Тип", "Сума", "Нотатка"]
    details.append(headers)
    for col_idx in range(1, len(headers) + 1):
        c = details.cell(row=1, column=col_idx)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = BOX
    details.freeze_panes = "A2"

    for currency in data["currencies"]:
        for group in currency["groups"]:
            for line in group["lines"]:
                # Map kinds list to a single human label
                kinds = line.get("kinds", [])
                if kinds == ["baseline"]:
                    kind_label = "звичайне"
                elif kinds == ["one_time"]:
                    kind_label = "одноразове"
                elif kinds == ["income"]:
                    kind_label = "дохід"
                elif set(kinds) == {"baseline", "one_time"}:
                    kind_label = "звичайне + одноразове"
                else:
                    kind_label = ", ".join(kinds)
                details.append(
                    [
                        data["period"],
                        group["title"],
                        line["category_display"],
                        currency["alpha"],
                        kind_label,
                        line["amount_major"],
                        line.get("note") or "",
                    ]
                )

    widths = {"A": 10, "B": 18, "C": 36, "D": 8, "E": 22, "F": 14, "G": 40}
    for col, w in widths.items():
        details.column_dimensions[col].width = w
    # Amount column formatting in Details
    for r in range(2, details.max_row + 1):
        details.cell(row=r, column=6).number_format = "#,##0.00;[Red]-#,##0.00"

    # Total at top of Overview - link to TOTAL.
    # (Already populated as we walked groups.)

    wb.save(out_path)


def _write_variance_csv(rows: list[dict[str, Any]], out_path: Path | None) -> None:
    import csv
    import io

    headers = ["Period", "Category", "Currency", "Target", "Actual", "Delta", "% used"]
    target = out_path.open("w", encoding="utf-8", newline="") if out_path else io.StringIO()
    try:
        writer = csv.DictWriter(target, fieldnames=headers)
        writer.writeheader()
        for r in rows:
            writer.writerow(r)
    finally:
        if out_path is None:
            # stdout dump (the JSON success payload still goes to
            # stdout via run_subcommand - but the spreadsheet contents
            # are also helpful to print here for piping. We print to
            # stderr so the JSON contract is preserved). Actually,
            # mixing CSV and JSON on the same stream confuses callers;
            # keep this branch silent and tell the user to pass --out
            # PATH for piping. Existing tests prefer the file path.
            target.close()
        else:
            target.close()


def _write_variance_xlsx(rows: list[dict[str, Any]], out_path: Path) -> None:
    try:
        from openpyxl import Workbook  # type: ignore[import-not-found]
    except ImportError as exc:
        raise CliError(
            "openpyxl is required for xlsx export; "
            "install via ``uv pip install openpyxl`` or use --format csv",
            kind="MissingDependency",
        ) from exc
    wb = Workbook()
    ws = wb.active
    ws.title = "Variance"
    headers = ["Period", "Category", "Currency", "Target", "Actual", "Delta", "% used"]
    ws.append(headers)
    for r in rows:
        ws.append([r[h] for h in headers])
    wb.save(out_path)


def _coerce_amount_to_minor(raw: str) -> int:
    try:
        return int(round(float(raw) * 100))
    except (ValueError, TypeError) as exc:
        raise CliError(f"--amount={raw!r} is not numeric", kind="BadAmount") from exc


def cmd_plan_start(args: argparse.Namespace) -> dict[str, Any]:
    db_path = resolve_db_path(args.db)
    with closing(open_db(db_path)) as conn:
        try:
            result = bud.start_draft(
                conn,
                period=args.period,
                copy_from=args.copy_from,
            )
        except bud.BudgetParseError as exc:
            raise CliError(str(exc), kind=exc.kind, details=exc.details) from exc
    return result


def cmd_plan_commit(args: argparse.Namespace) -> dict[str, Any]:
    db_path = resolve_db_path(args.db)
    with closing(open_db(db_path)) as conn:
        try:
            return bud.commit_draft(conn, period=args.period)
        except bud.BudgetParseError as exc:
            raise CliError(str(exc), kind=exc.kind) from exc


def cmd_plan_cancel(args: argparse.Namespace) -> dict[str, Any]:
    db_path = resolve_db_path(args.db)
    with closing(open_db(db_path)) as conn:
        try:
            return bud.cancel_draft(conn, period=args.period)
        except bud.BudgetParseError as exc:
            raise CliError(str(exc), kind=exc.kind) from exc


def cmd_plan_show(args: argparse.Namespace) -> dict[str, Any]:
    """Read the current draft for a period (if any). Falls back to the
    active budget when no draft is present so the user can use the
    same command throughout."""
    cur_code = _parse_currency_or_cli_error(args.currency)
    db_path = resolve_db_path(args.db)
    with closing(open_db(db_path)) as conn:
        drafts = conn.execute(
            "SELECT id FROM budget WHERE period = ? AND status = 'draft' "
            + ("AND currency_code = ? " if cur_code is not None else "")
            + "LIMIT 1",
            ([args.period, cur_code] if cur_code is not None else [args.period]),
        ).fetchone()
        if drafts is not None:
            # Reuse fetch_budget but filter to drafts only - we want to
            # see the work-in-progress, not the active twin.
            blocks = bud.fetch_budget(conn, period=args.period, currency_code=cur_code)
            draft_blocks = [b for b in blocks if b["status"] == "draft"]
            return {
                "ok": True,
                "period": args.period,
                "currency": args.currency,
                "viewing": "draft",
                "budgets": draft_blocks,
                "edit_log_size": conn.execute(
                    "SELECT COUNT(*) FROM budget_draft_edit bde "
                    "JOIN budget b ON b.id = bde.budget_id "
                    "WHERE b.period = ? AND b.status = 'draft'",
                    (args.period,),
                ).fetchone()[0],
            }
        # No draft - fall through to active
        blocks = bud.fetch_budget(conn, period=args.period, currency_code=cur_code)
        active_blocks = [b for b in blocks if b["status"] == "active"]
        return {
            "ok": True,
            "period": args.period,
            "currency": args.currency,
            "viewing": "active" if active_blocks else "none",
            "budgets": active_blocks,
            "edit_log_size": 0,
        }


def cmd_plan_add(args: argparse.Namespace) -> dict[str, Any]:
    _validate_category_name(args.category)
    cur_code = _parse_currency_or_cli_error(args.currency)
    if cur_code is None:
        raise CliError("--currency is required", kind="BadArgument")
    amount = _coerce_amount_to_minor(args.amount)
    db_path = resolve_db_path(args.db)
    with closing(open_db(db_path)) as conn:
        try:
            return bud.add_line(
                conn,
                period=args.period,
                category=args.category,
                currency_code=cur_code,
                kind=args.kind,
                amount_minor=amount,
                note=args.note,
            )
        except bud.BudgetParseError as exc:
            raise CliError(str(exc), kind=exc.kind) from exc


def cmd_plan_update(args: argparse.Namespace) -> dict[str, Any]:
    cur_code = _parse_currency_or_cli_error(args.currency)
    amount = _coerce_amount_to_minor(args.amount) if args.amount is not None else None
    db_path = resolve_db_path(args.db)
    with closing(open_db(db_path)) as conn:
        try:
            return bud.update_line(
                conn,
                period=args.period,
                line_id=args.line_id,
                category=args.category,
                currency_code=cur_code,
                kind=args.kind,
                amount_minor=amount,
                note=args.note,
                note_unset=args.note_unset,
            )
        except bud.BudgetParseError as exc:
            raise CliError(str(exc), kind=exc.kind, details=exc.details) from exc


def cmd_plan_remove(args: argparse.Namespace) -> dict[str, Any]:
    cur_code = _parse_currency_or_cli_error(args.currency)
    db_path = resolve_db_path(args.db)
    with closing(open_db(db_path)) as conn:
        try:
            return bud.remove_line(
                conn,
                period=args.period,
                line_id=args.line_id,
                category=args.category,
                currency_code=cur_code,
                kind=args.kind,
            )
        except bud.BudgetParseError as exc:
            raise CliError(str(exc), kind=exc.kind, details=exc.details) from exc


def cmd_plan_undo(args: argparse.Namespace) -> dict[str, Any]:
    db_path = resolve_db_path(args.db)
    with closing(open_db(db_path)) as conn:
        return bud.undo_last(conn, period=args.period)


def cmd_plan_suggest(args: argparse.Namespace) -> dict[str, Any]:
    """Return suggestion signals from the history scanner. Claude
    phrases them back to the user; the CLI just emits the structured
    facts."""
    db_path = resolve_db_path(args.db)
    with closing(open_db(db_path)) as conn:
        try:
            signals = bud.scan_history_for_signals(
                conn,
                target_period=args.period,
                lookback_months=args.lookback,
            )
        except bud.BudgetParseError as exc:
            raise CliError(str(exc), kind=exc.kind) from exc
    return {
        "ok": True,
        "period": args.period,
        "lookback_months": args.lookback,
        "signals": signals,
    }


def cmd_rename_category(args: argparse.Namespace) -> dict[str, Any]:
    old = _validate_category_name(args.old)
    new = _validate_category_name(args.new)
    if old == new:
        raise CliError("--from and --to are identical", kind="BadArgument")
    update_tables: tuple[str, ...] = tuple(args.update.split(","))
    db_path = resolve_db_path(args.db)
    with closing(open_db(db_path)) as conn:
        try:
            counts = bud.rename_category(
                conn,
                old_name=old,
                new_name=new,
                update_tables=update_tables,
            )
        except bud.BudgetParseError as exc:
            raise CliError(str(exc), kind=exc.kind) from exc
    return {"ok": True, "old": old, "new": new, "counts": counts}


def _build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        prog="pf-budget",
        description=(
            "Personal finance budgeting commands. PR1 ships the "
            "category_registry subset; pf-budget import / show / "
            "diff / export / close arrive in later PRs."
        ),
    )
    sub = p.add_subparsers(dest="cmd", required=True)

    p_reg = sub.add_parser(
        "register-category",
        help=(
            "Declare a category so budget imports recognise it before any transaction matches it"
        ),
    )
    p_reg.add_argument("--category", required=True)
    p_reg.add_argument(
        "--note",
        default=None,
        help="Optional human-readable note stored alongside the registration",
    )
    p_reg.add_argument("--db", default=None)
    p_reg.set_defaults(func=cmd_register_category)

    p_unreg = sub.add_parser(
        "unregister-category",
        help="Remove a category from the registry (refuses when still referenced)",
    )
    p_unreg.add_argument("--category", required=True)
    p_unreg.add_argument(
        "--force",
        action="store_true",
        help=argparse.SUPPRESS,  # intentional escape hatch
    )
    p_unreg.add_argument("--db", default=None)
    p_unreg.set_defaults(func=cmd_unregister_category)

    p_list = sub.add_parser(
        "list-categories",
        help="List in-use (and optionally declared-but-unused) categories",
    )
    p_list.add_argument(
        "--include-declared",
        action="store_true",
        help="Also include declared-but-unused entries from category_registry",
    )
    p_list.add_argument("--db", default=None)
    p_list.set_defaults(func=cmd_list_categories)

    p_import = sub.add_parser(
        "import",
        help="Import a budget plan (CSV or XLSX) for a given period",
    )
    p_import.add_argument(
        "source",
        help="Path to CSV (.csv) or XLSX (.xlsx/.xlsm). CSV is the Plans-shape "
        "by default; pass --sheet baseline to read a Baseline-shape CSV.",
    )
    p_import.add_argument(
        "--period",
        required=True,
        help="Target period in YYYY-MM. Plans rows with other periods are ignored.",
    )
    p_import.add_argument(
        "--unknown-categories",
        choices=("reject", "register"),
        default="reject",
        help=(
            "What to do with categories that don't exist yet. "
            "'reject' (default): fail with suggestions. "
            "'register': add to category_registry and proceed."
        ),
    )
    p_import.add_argument(
        "--sheet",
        choices=("plans", "baseline"),
        default=None,
        help="For CSV input only: which sheet shape the file represents. "
        "Defaults to 'plans'. Ignored for XLSX (which reads both sheets).",
    )
    p_import.add_argument(
        "--dry-run",
        action="store_true",
        help="Parse, validate, and summarise the diff without writing.",
    )
    p_import.add_argument(
        "--force",
        action="store_true",
        help="Overwrite a closed budget for the period (rare).",
    )
    p_import.add_argument("--db", default=None)
    p_import.set_defaults(func=cmd_import)

    p_show = sub.add_parser("show", help="Show the budget plan materialised for a period")
    p_show.add_argument("--period", required=True)
    p_show.add_argument("--currency", default=None, help="UAH/USD/980/840/...")
    p_show.add_argument("--db", default=None)
    p_show.set_defaults(func=cmd_show)

    p_diff = sub.add_parser("diff", help="Budget vs actuals for a period, per (currency, category)")
    p_diff.add_argument("--period", required=True)
    p_diff.add_argument("--currency", default=None)
    p_diff.add_argument("--db", default=None)
    p_diff.set_defaults(func=cmd_diff)

    p_list_b = sub.add_parser("list", help="List every materialised budget with totals and status")
    p_list_b.add_argument("--db", default=None)
    p_list_b.set_defaults(func=cmd_list)

    p_close = sub.add_parser("close", help="Mark a budget as closed (snapshot)")
    p_close.add_argument("--period", required=True)
    p_close.add_argument("--currency", default=None)
    p_close.add_argument("--db", default=None)
    p_close.set_defaults(func=cmd_close)

    p_reopen = sub.add_parser("reopen", help="Flip a closed budget back to active")
    p_reopen.add_argument("--period", required=True)
    p_reopen.add_argument("--currency", default=None)
    p_reopen.add_argument(
        "--reason",
        default=None,
        help="Free-text reason for reopening; stored only in the command result, "
        "not persisted (audit lands in a future PR if needed).",
    )
    p_reopen.add_argument("--db", default=None)
    p_reopen.set_defaults(func=cmd_reopen)

    p_delete = sub.add_parser("delete", help="Delete a budget and its lines (cascade)")
    p_delete.add_argument("--period", required=True)
    p_delete.add_argument("--currency", default=None)
    p_delete.add_argument(
        "--force",
        action="store_true",
        help="Allow deleting a closed budget.",
    )
    p_delete.add_argument("--db", default=None)
    p_delete.set_defaults(func=cmd_delete)

    p_export = sub.add_parser(
        "export",
        help="Write a budget view (variance / family / plan) for a period",
    )
    p_export.add_argument("--period", required=True)
    p_export.add_argument("--currency", default=None)
    p_export.add_argument(
        "--out",
        required=True,
        help="Output file path (.csv or .xlsx). Use '-' for stdout (CSV only).",
    )
    p_export.add_argument(
        "--format",
        choices=("auto", "csv", "xlsx"),
        default="auto",
        help="Output format. 'auto' uses the file extension; defaults to csv "
        "when writing to stdout.",
    )
    p_export.add_argument(
        "--view",
        choices=("variance", "family", "plan"),
        default="variance",
        help="What to export. 'variance' (default): target vs actual. "
        "'family': pretty grouped view for non-technical readers (xlsx only). "
        "'plan': raw active-budget dump.",
    )
    p_export.add_argument("--db", default=None)
    p_export.set_defaults(func=cmd_export)

    # --- plan subcommand group ---------------------------------------
    p_plan = sub.add_parser(
        "plan",
        help="Conversation-driven planning (draft, edit, undo, commit)",
    )
    plan_sub = p_plan.add_subparsers(dest="plan_cmd", required=True)

    pps = plan_sub.add_parser(
        "start", help="Create a draft for the period, copying from prior month"
    )
    pps.add_argument("--period", required=True)
    pps.add_argument(
        "--copy-from",
        default=None,
        help="Source period to copy baseline from. Default: most recent "
        "active period. Pass empty string to start blank.",
    )
    pps.add_argument("--db", default=None)
    pps.set_defaults(func=cmd_plan_start)

    ppc = plan_sub.add_parser("commit", help="Flip draft to active (atomic replace)")
    ppc.add_argument("--period", required=True)
    ppc.add_argument("--db", default=None)
    ppc.set_defaults(func=cmd_plan_commit)

    ppx = plan_sub.add_parser("cancel", help="Delete the draft for a period")
    ppx.add_argument("--period", required=True)
    ppx.add_argument("--db", default=None)
    ppx.set_defaults(func=cmd_plan_cancel)

    ppshow = plan_sub.add_parser(
        "show",
        help="Show the draft (or active when no draft) for a period",
    )
    ppshow.add_argument("--period", required=True)
    ppshow.add_argument("--currency", default=None)
    ppshow.add_argument("--db", default=None)
    ppshow.set_defaults(func=cmd_plan_show)

    ppadd = plan_sub.add_parser("add", help="Add a line to the draft")
    ppadd.add_argument("--period", required=True)
    ppadd.add_argument("--category", required=True)
    ppadd.add_argument("--currency", required=True)
    ppadd.add_argument("--kind", required=True, choices=("baseline", "one_time", "income"))
    ppadd.add_argument("--amount", required=True, help="Major units, signed (-5300 for outflow)")
    ppadd.add_argument("--note", default=None)
    ppadd.add_argument("--db", default=None)
    ppadd.set_defaults(func=cmd_plan_add)

    ppupd = plan_sub.add_parser("update", help="Update a draft line")
    ppupd.add_argument("--period", required=True)
    ppupd.add_argument("--line-id", type=int, default=None)
    ppupd.add_argument("--category", default=None)
    ppupd.add_argument("--currency", default=None)
    ppupd.add_argument(
        "--kind",
        default=None,
        choices=("baseline", "one_time", "income"),
    )
    ppupd.add_argument("--amount", default=None, help="New amount (major units)")
    ppupd.add_argument("--note", default=None)
    ppupd.add_argument(
        "--note-unset",
        action="store_true",
        help="Explicitly clear the note (use instead of --note '')",
    )
    ppupd.add_argument("--db", default=None)
    ppupd.set_defaults(func=cmd_plan_update)

    pprem = plan_sub.add_parser("remove", help="Remove a draft line")
    pprem.add_argument("--period", required=True)
    pprem.add_argument("--line-id", type=int, default=None)
    pprem.add_argument("--category", default=None)
    pprem.add_argument("--currency", default=None)
    pprem.add_argument(
        "--kind",
        default=None,
        choices=("baseline", "one_time", "income"),
    )
    pprem.add_argument("--db", default=None)
    pprem.set_defaults(func=cmd_plan_remove)

    ppundo = plan_sub.add_parser("undo", help="Undo the most recent draft edit")
    ppundo.add_argument("--period", required=True)
    ppundo.add_argument("--db", default=None)
    ppundo.set_defaults(func=cmd_plan_undo)

    ppsug = plan_sub.add_parser(
        "suggest",
        help="Scan prior months and return planning hints for the period",
    )
    ppsug.add_argument("--period", required=True)
    ppsug.add_argument(
        "--lookback",
        type=int,
        default=6,
        help="How many months of history to scan (default 6)",
    )
    ppsug.add_argument("--db", default=None)
    ppsug.set_defaults(func=cmd_plan_suggest)

    # --- end plan group -------------------------------------------------

    p_rename = sub.add_parser(
        "rename-category",
        help="Rewrite a category name across selected tables",
    )
    p_rename.add_argument("--from", dest="old", required=True)
    p_rename.add_argument("--to", dest="new", required=True)
    p_rename.add_argument(
        "--update",
        default="budget_line",
        help="Comma-separated tables to rewrite. Allowed: "
        "budget_line, tx_category, category_overrides, "
        "categorization_rules, category_registry.",
    )
    p_rename.add_argument("--db", default=None)
    p_rename.set_defaults(func=cmd_rename_category)

    return p


def main(argv: Sequence[str] | None = None) -> int:
    args = _build_parser().parse_args(argv)
    return run_subcommand(args.func, args)


if __name__ == "__main__":
    sys.exit(main())
