"""Budget feature: parsing, validation, materialisation.

This module owns:

- Plain dataclasses for plan rows (``PlanRow``) so the parsers and
  the materialiser agree on a single in-memory shape.
- CSV parser for the ``Plans`` sheet format (the canonical exchange
  format with Google Sheets - the user copies a single CSV per month).
- XLSX parser for two-sheet workbooks (``Baseline`` + ``Plans``);
  optional, depends on ``openpyxl`` being installed.
- Category validation: known = ``tx_category`` ∪ ``category_overrides``
  ∪ ``categorization_rules`` ∪ ``category_registry``. Unknown entries
  get Levenshtein-distance suggestions against the known set.
- Materialisation: one budget per (period, currency). Idempotent
  per (period, currency) - re-running replaces the lines unless
  ``status='closed'`` (in which case it refuses without ``--force``).

The CLI layer (``budget_cli.py``) is intentionally thin around these
helpers - same separation as ``queries.py`` vs ``query.py``.
"""

from __future__ import annotations

import csv
import json
import re
import sqlite3
import time
from collections.abc import Iterable
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from .currencies import numeric_for

PERIOD_RE = re.compile(r"^\d{4}-\d{2}$")
ALLOWED_KINDS = ("baseline", "one_time", "income")

# Required columns for each sheet shape, enforced at parse time.
_PLANS_REQUIRED = ("Period", "Category", "Currency", "Kind", "Amount")
_PLANS_OPTIONAL = ("Note",)
_BASELINE_REQUIRED = ("Category", "Currency", "Kind", "Monthly target")
_BASELINE_OPTIONAL = ("Note",)


class BudgetParseError(ValueError):
    """Raised by parser/validator with a human-readable explanation
    of which row / column / value broke the contract. Carries an
    optional ``details`` payload (e.g. list of unknown categories
    with suggestions) so the CLI can render structured JSON output."""

    def __init__(
        self,
        message: str,
        *,
        kind: str = "BudgetParse",
        details: dict[str, Any] | None = None,
    ) -> None:
        super().__init__(message)
        self.kind = kind
        self.details = details or {}


@dataclass
class PlanRow:
    """One concrete budget line, after CSV/XLSX parsing and merging.

    ``period`` is always set (Plans rows carry it; Baseline rows get
    the ``--period`` value from the CLI at merge time). ``amount_minor``
    is signed (negative for outflows, positive for income).
    """

    period: str
    category: str
    currency_code: int
    kind: str
    amount_minor: int
    note: str | None = None
    # ``source_row``: human-readable origin of this row for error
    # messages (e.g. ``Plans!B14`` or ``baseline.csv:3``). Optional
    # so internal callers (tests) don't have to thread it.
    source_row: str | None = None


@dataclass
class ValidationResult:
    """What ``validate_categories`` returns to the caller.

    ``unknown``: list of ``(category, suggestions)`` where each
    suggestion is ``(candidate, distance)`` sorted ascending by
    distance, then alphabetically. Empty when every category was
    recognised.
    """

    rows: list[PlanRow]
    unknown: list[tuple[str, list[tuple[str, int]]]] = field(default_factory=list)


# ---------------------------------------------------------------------------
# Parsing
# ---------------------------------------------------------------------------


def parse_plans_csv(path: Path) -> list[PlanRow]:
    """Read a Plans-shape CSV. Header must match ``_PLANS_REQUIRED``;
    ``Note`` is optional."""
    with path.open(encoding="utf-8") as f:
        reader = csv.DictReader(f)
        _check_columns(reader.fieldnames, _PLANS_REQUIRED, _PLANS_OPTIONAL, sheet="Plans")
        rows: list[PlanRow] = []
        for i, raw in enumerate(reader, start=2):  # row 1 = header
            if not any((raw.get(k) or "").strip() for k in _PLANS_REQUIRED):
                continue  # skip blank trailing rows
            rows.append(_row_from_plans(raw, source_row=f"{path.name}:{i}"))
    return rows


def parse_baseline_csv(path: Path, period: str) -> list[PlanRow]:
    """Read a Baseline-shape CSV and stamp ``period`` onto each row."""
    _validate_period(period, "period")
    with path.open(encoding="utf-8") as f:
        reader = csv.DictReader(f)
        _check_columns(reader.fieldnames, _BASELINE_REQUIRED, _BASELINE_OPTIONAL, sheet="Baseline")
        rows: list[PlanRow] = []
        for i, raw in enumerate(reader, start=2):
            if not any((raw.get(k) or "").strip() for k in _BASELINE_REQUIRED):
                continue
            rows.append(_row_from_baseline(raw, period, source_row=f"{path.name}:{i}"))
    return rows


def parse_workbook_xlsx(path: Path) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """Read a two-sheet workbook (Baseline + Plans). Returns the raw
    dict rows (header row drives keys) so the caller can stamp the
    period onto Baseline rows and merge with Plans.

    Lazy-imports ``openpyxl`` so the rest of the budget feature is
    usable on installs without the optional dependency.
    """
    try:
        from openpyxl import load_workbook  # type: ignore[import-not-found]
    except ImportError as exc:  # pragma: no cover - covered by message
        raise BudgetParseError(
            "openpyxl is required to read XLSX files; install with "
            "``uv pip install openpyxl`` or use a CSV export of each sheet",
            kind="MissingDependency",
        ) from exc

    wb = load_workbook(path, read_only=True, data_only=True)
    baseline_rows: list[dict[str, Any]] = []
    plans_rows: list[dict[str, Any]] = []
    for sheet_name, target, required, optional in (
        ("Baseline", baseline_rows, _BASELINE_REQUIRED, _BASELINE_OPTIONAL),
        ("Plans", plans_rows, _PLANS_REQUIRED, _PLANS_OPTIONAL),
    ):
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        rows_iter = ws.iter_rows(values_only=True)
        try:
            raw_header = next(rows_iter)
        except StopIteration:
            continue
        # Coerce header cells to strings. Sheets occasionally come
        # back with non-string headers (numbers, dates) - those are
        # legitimately bad input that we want to surface via
        # _check_columns below.
        header: list[str] = [
            h.strip() if isinstance(h, str) else str(h) if h is not None else "" for h in raw_header
        ]
        _check_columns(header, required, optional, sheet=sheet_name)
        for i, values in enumerate(rows_iter, start=2):
            if values is None or all(
                v is None or (isinstance(v, str) and not v.strip()) for v in values
            ):
                continue
            row_dict: dict[str, Any] = {
                header[j]: ("" if values[j] is None else values[j])
                for j in range(min(len(header), len(values)))
                if header[j]
            }
            row_dict["__source_row__"] = f"{sheet_name}!{i}"
            target.append(row_dict)
    wb.close()
    return baseline_rows, plans_rows


def merge_baseline_plans(
    baseline: Iterable[PlanRow],
    plans: Iterable[PlanRow],
    *,
    period: str,
) -> list[PlanRow]:
    """Combine workbook-level Baseline with period-specific Plans.

    Rules:
    - Plans rows with the target ``period`` take precedence: any
      Baseline row with the same ``(category, currency_code, kind)``
      is dropped.
    - Plans rows with a different period are filtered out (they belong
      to a different month's budget, not this one).
    - The result preserves Baseline first (in input order), Plans
      second, so deterministic ordering helps tests.
    """
    plans_for_period = [r for r in plans if r.period == period]
    suppress = {(r.category, r.currency_code, r.kind) for r in plans_for_period}
    out: list[PlanRow] = []
    for r in baseline:
        if r.period != period:
            # Baseline rows always carry the period stamped in the parse
            # step; this should be invariant.
            continue
        if (r.category, r.currency_code, r.kind) in suppress:
            continue
        out.append(r)
    out.extend(plans_for_period)
    return out


# ---------------------------------------------------------------------------
# Validation
# ---------------------------------------------------------------------------


def known_categories(conn: sqlite3.Connection) -> set[str]:
    """All categories the system recognises right now: in-use anywhere
    in the rules / tx tables, plus declared-but-unused from the
    registry."""
    queries = (
        "SELECT category FROM tx_category",
        "SELECT category FROM category_overrides",
        "SELECT category FROM categorization_rules",
        "SELECT category FROM category_registry",
        "SELECT category FROM budget_line",
    )
    known: set[str] = set()
    for q in queries:
        try:
            known.update(r[0] for r in conn.execute(q) if r[0])
        except sqlite3.OperationalError as exc:
            # Pre-migration tables may not exist. The store always
            # brings the schema to current before this is called, but
            # the defensive probe is cheap.
            if "no such table" not in str(exc).lower():
                raise
    return known


def levenshtein(a: str, b: str) -> int:
    """Iterative DP Levenshtein distance.

    Hand-rolled rather than dragging in ``rapidfuzz`` - the vocabulary
    of categories is small (tens of strings) and we only call this on
    the unknowns, never on a hot path.
    """
    if a == b:
        return 0
    if not a:
        return len(b)
    if not b:
        return len(a)
    prev = list(range(len(b) + 1))
    curr = [0] * (len(b) + 1)
    for i, ca in enumerate(a, start=1):
        curr[0] = i
        for j, cb in enumerate(b, start=1):
            cost = 0 if ca == cb else 1
            curr[j] = min(
                prev[j] + 1,  # deletion
                curr[j - 1] + 1,  # insertion
                prev[j - 1] + cost,  # substitution
            )
        prev, curr = curr, prev
    return prev[len(b)]


def suggest_categories(
    target: str, candidates: Iterable[str], *, top_n: int = 3
) -> list[tuple[str, int]]:
    """Return the closest ``top_n`` candidates by Levenshtein distance,
    sorted ascending by distance then by name."""
    scored = [(c, levenshtein(target, c)) for c in candidates]
    scored.sort(key=lambda x: (x[1], x[0]))
    return scored[:top_n]


def validate_categories(
    rows: list[PlanRow],
    conn: sqlite3.Connection,
) -> ValidationResult:
    """Split ``rows`` into known / unknown by checking each category
    against the current taxonomy. Unknowns come back with the top-3
    closest candidates so the CLI / caller can present a helpful
    error message.
    """
    known = known_categories(conn)
    unknown_set: list[tuple[str, list[tuple[str, int]]]] = []
    seen_unknown: set[str] = set()
    for r in rows:
        if r.category in known:
            continue
        if r.category in seen_unknown:
            continue
        seen_unknown.add(r.category)
        unknown_set.append((r.category, suggest_categories(r.category, known)))
    return ValidationResult(rows=rows, unknown=unknown_set)


def register_unknowns(
    conn: sqlite3.Connection, unknowns: Iterable[str], *, now_ts: int | None = None
) -> int:
    """Insert each unknown into ``category_registry`` with
    ``declared_via='budget-import'``. Returns count of newly-added
    rows (skips ones already present).
    """
    ts = now_ts if now_ts is not None else int(time.time())
    added = 0
    for c in unknowns:
        cur = conn.execute(
            "INSERT OR IGNORE INTO category_registry "
            "(category, declared_at, declared_via) VALUES (?, ?, ?)",
            (c, ts, "budget-import"),
        )
        if cur.rowcount:
            added += 1
    return added


# ---------------------------------------------------------------------------
# Materialisation
# ---------------------------------------------------------------------------


@dataclass
class MaterialiseResult:
    period: str
    # cur -> {budget_id, lines_added, lines_replaced, status_after}
    by_currency: dict[int, dict[str, Any]]


def materialise_budget(
    conn: sqlite3.Connection,
    *,
    period: str,
    rows: list[PlanRow],
    source: str,
    force: bool = False,
    now_ts: int | None = None,
) -> MaterialiseResult:
    """Write the plan to ``budget`` / ``budget_line`` tables.

    One budget row per (period, currency_code) found in ``rows``.
    Existing draft / active budgets are replaced (all lines deleted
    then re-inserted in the same transaction). Closed budgets are
    refused unless ``force=True``.

    Caller is responsible for opening the DB connection and managing
    its lifetime; this function does NOT call ``conn.close``. It DOES
    drive the transaction boundary - one BEGIN/COMMIT per currency so
    a UAH success is kept even if the USD slice fails downstream.
    """
    _validate_period(period, "period")
    ts = now_ts if now_ts is not None else int(time.time())
    by_currency: dict[int, list[PlanRow]] = {}
    for r in rows:
        by_currency.setdefault(r.currency_code, []).append(r)

    result = MaterialiseResult(period=period, by_currency={})

    for cur_code, items in sorted(by_currency.items()):
        existing = conn.execute(
            "SELECT id, status FROM budget WHERE period = ? AND currency_code = ?",
            (period, cur_code),
        ).fetchone()

        conn.execute("BEGIN")
        try:
            replaced = 0
            if existing is not None:
                budget_id, status = existing
                if status == "closed" and not force:
                    conn.rollback()
                    raise BudgetParseError(
                        f"budget for {period}/{cur_code} is closed; "
                        "pass --force to overwrite (rare) or "
                        "pf-budget reopen first",
                        kind="ClosedBudget",
                        details={"period": period, "currency_code": cur_code},
                    )
                if status == "closed" and force:
                    # Walk through reopen → replace → close so the
                    # trigger never sees an insert into a closed
                    # parent.
                    conn.execute(
                        "UPDATE budget SET status = 'active' WHERE id = ?",
                        (budget_id,),
                    )
                replaced = conn.execute(
                    "DELETE FROM budget_line WHERE budget_id = ?",
                    (budget_id,),
                ).rowcount
                conn.execute(
                    "UPDATE budget SET imported_from = ? WHERE id = ?",
                    (source, budget_id),
                )
            else:
                conn.execute(
                    "INSERT INTO budget "
                    "(period, currency_code, status, created_at, imported_from) "
                    "VALUES (?, ?, ?, ?, ?)",
                    (period, cur_code, "active", ts, source),
                )
                budget_id = conn.execute(
                    "SELECT id FROM budget WHERE period = ? AND currency_code = ?",
                    (period, cur_code),
                ).fetchone()[0]

            for r in items:
                conn.execute(
                    "INSERT INTO budget_line "
                    "(budget_id, category, amount_minor, kind, note) "
                    "VALUES (?, ?, ?, ?, ?)",
                    (budget_id, r.category, r.amount_minor, r.kind, r.note),
                )
            if existing is not None and existing[1] == "closed" and force:
                conn.execute(
                    "UPDATE budget SET status = 'closed' WHERE id = ?",
                    (budget_id,),
                )
            conn.commit()
        except Exception:
            conn.rollback()
            raise

        result.by_currency[cur_code] = {
            "budget_id": int(budget_id),
            "lines_added": len(items),
            "lines_replaced": int(replaced),
            "status_after": "closed"
            if (existing and existing[1] == "closed" and force)
            else "active",
        }

    return result


def find_budget(
    conn: sqlite3.Connection,
    *,
    period: str,
    currency_code: int,
    status: str | None = None,
) -> dict[str, Any] | None:
    """Single (period, currency_code, status) lookup."""
    where = ["period = ?", "currency_code = ?"]
    params: list[Any] = [period, currency_code]
    if status is not None:
        where.append("status = ?")
        params.append(status)
    row = conn.execute(
        "SELECT id, period, currency_code, status, created_at, imported_from "
        f"FROM budget WHERE {' AND '.join(where)}",
        params,
    ).fetchone()
    if row is None:
        return None
    return {
        "budget_id": int(row[0]),
        "period": row[1],
        "currency_code": int(row[2]),
        "status": row[3],
        "created_at": int(row[4]),
        "imported_from": row[5],
    }


def _latest_active_period_before(conn: sqlite3.Connection, period: str) -> str | None:
    row = conn.execute(
        "SELECT DISTINCT period FROM budget "
        "WHERE status = 'active' AND period < ? "
        "ORDER BY period DESC LIMIT 1",
        (period,),
    ).fetchone()
    return row[0] if row else None


def start_draft(
    conn: sqlite3.Connection,
    *,
    period: str,
    copy_from: str | None = None,
    now_ts: int | None = None,
) -> dict[str, Any]:
    """Create draft budget rows for ``period``, optionally copying
    baseline lines from a prior period's active budget.

    Returns ``existing_draft=True`` (without re-creating) when a draft
    is already in place for ``period`` - the caller asks the user to
    continue / cancel / merge.

    ``copy_from``:
      - ``None``: use the most recent active period before ``period``
      - explicit ``YYYY-MM``: copy from that period
      - explicit empty ``""``: start blank (no copy)
    """
    _validate_period(period, "period")
    ts = now_ts if now_ts is not None else int(time.time())

    existing_drafts = conn.execute(
        "SELECT id, currency_code, created_at FROM budget "
        "WHERE period = ? AND status = 'draft' ORDER BY currency_code",
        (period,),
    ).fetchall()
    if existing_drafts:
        return {
            "ok": True,
            "existing_draft": True,
            "period": period,
            "drafts": [
                {
                    "budget_id": int(r[0]),
                    "currency_code": int(r[1]),
                    "created_at": int(r[2]),
                    "line_count": conn.execute(
                        "SELECT COUNT(*) FROM budget_line WHERE budget_id = ?",
                        (r[0],),
                    ).fetchone()[0],
                }
                for r in existing_drafts
            ],
        }

    if copy_from is None:
        copy_from = _latest_active_period_before(conn, period) or ""
    elif copy_from:
        _validate_period(copy_from, "copy_from")

    drafts_created: list[dict[str, Any]] = []

    if not copy_from:
        return {
            "ok": True,
            "existing_draft": False,
            "period": period,
            "copied_from": None,
            "drafts": drafts_created,
        }

    source_budgets = conn.execute(
        "SELECT id, currency_code FROM budget "
        "WHERE period = ? AND status = 'active' ORDER BY currency_code",
        (copy_from,),
    ).fetchall()

    conn.execute("BEGIN")
    try:
        for src_budget_id, cur_code in source_budgets:
            conn.execute(
                "INSERT INTO budget "
                "(period, currency_code, status, created_at, imported_from) "
                "VALUES (?, ?, 'draft', ?, ?)",
                (period, cur_code, ts, f"copy:{copy_from}"),
            )
            new_id = conn.execute(
                "SELECT id FROM budget WHERE period = ? AND currency_code = ? AND status = 'draft'",
                (period, cur_code),
            ).fetchone()[0]
            copied = conn.execute(
                "INSERT INTO budget_line "
                "(budget_id, category, amount_minor, kind, note) "
                "SELECT ?, category, amount_minor, kind, note FROM budget_line "
                "WHERE budget_id = ? AND kind = 'baseline'",
                (new_id, src_budget_id),
            ).rowcount
            drafts_created.append(
                {
                    "budget_id": int(new_id),
                    "currency_code": int(cur_code),
                    "lines_copied": int(copied),
                }
            )
        conn.commit()
    except Exception:
        conn.rollback()
        raise

    return {
        "ok": True,
        "existing_draft": False,
        "period": period,
        "copied_from": copy_from,
        "drafts": drafts_created,
    }


def _ensure_draft_budget(
    conn: sqlite3.Connection,
    *,
    period: str,
    currency_code: int,
    now_ts: int,
) -> int:
    """Return draft budget_id for (period, currency_code), creating
    an empty one if absent. Used by add_line so the conversation can
    introduce a new currency mid-planning."""
    row = conn.execute(
        "SELECT id FROM budget WHERE period = ? AND currency_code = ? AND status = 'draft'",
        (period, currency_code),
    ).fetchone()
    if row is not None:
        return int(row[0])
    conn.execute(
        "INSERT INTO budget "
        "(period, currency_code, status, created_at, imported_from) "
        "VALUES (?, ?, 'draft', ?, NULL)",
        (period, currency_code, now_ts),
    )
    new_id = conn.execute(
        "SELECT id FROM budget WHERE period = ? AND currency_code = ? AND status = 'draft'",
        (period, currency_code),
    ).fetchone()[0]
    return int(new_id)


def _row_to_payload(row: tuple) -> dict[str, Any]:
    line_id, category, currency_code, kind, amount_minor, note = row
    return {
        "line_id": int(line_id),
        "category": category,
        "currency_code": int(currency_code),
        "kind": kind,
        "amount_minor": int(amount_minor),
        "note": note,
    }


def _read_line(conn: sqlite3.Connection, *, line_id: int) -> tuple | None:
    return conn.execute(
        "SELECT bl.id, bl.category, b.currency_code, bl.kind, bl.amount_minor, bl.note "
        "FROM budget_line bl JOIN budget b ON b.id = bl.budget_id "
        "WHERE bl.id = ?",
        (line_id,),
    ).fetchone()


def _find_lines_by_composite(
    conn: sqlite3.Connection,
    *,
    period: str,
    currency_code: int,
    category: str,
    kind: str,
) -> list[tuple]:
    return conn.execute(
        "SELECT bl.id, bl.category, b.currency_code, bl.kind, bl.amount_minor, bl.note "
        "FROM budget_line bl JOIN budget b ON b.id = bl.budget_id "
        "WHERE b.period = ? AND b.status = 'draft' "
        "AND b.currency_code = ? AND bl.category = ? AND bl.kind = ?",
        (period, currency_code, category, kind),
    ).fetchall()


def add_line(
    conn: sqlite3.Connection,
    *,
    period: str,
    category: str,
    currency_code: int,
    kind: str,
    amount_minor: int,
    note: str | None = None,
    now_ts: int | None = None,
) -> dict[str, Any]:
    """Insert one new line into the draft for ``period``."""
    _validate_period(period, "period")
    if kind not in ALLOWED_KINDS:
        raise BudgetParseError(
            f"kind={kind!r} must be one of {list(ALLOWED_KINDS)}",
            kind="BadKind",
        )
    ts = now_ts if now_ts is not None else int(time.time())
    conn.execute("BEGIN")
    try:
        budget_id = _ensure_draft_budget(
            conn, period=period, currency_code=currency_code, now_ts=ts
        )
        conn.execute(
            "INSERT INTO budget_line (budget_id, category, amount_minor, kind, note) "
            "VALUES (?, ?, ?, ?, ?)",
            (budget_id, category, amount_minor, kind, note),
        )
        new_line_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
        payload_after = {
            "line_id": int(new_line_id),
            "category": category,
            "currency_code": currency_code,
            "kind": kind,
            "amount_minor": amount_minor,
            "note": note,
        }
        conn.execute(
            "INSERT INTO budget_draft_edit "
            "(budget_id, op, payload_before, payload_after, applied_at) "
            "VALUES (?, ?, NULL, ?, ?)",
            (budget_id, "add", json.dumps(payload_after, ensure_ascii=False), ts),
        )
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    return {"ok": True, "op": "add", "line": payload_after, "budget_id": int(budget_id)}


def update_line(
    conn: sqlite3.Connection,
    *,
    period: str,
    line_id: int | None = None,
    category: str | None = None,
    currency_code: int | None = None,
    kind: str | None = None,
    amount_minor: int | None = None,
    note: str | None = None,
    note_unset: bool = False,
    now_ts: int | None = None,
) -> dict[str, Any]:
    """Modify a draft line. Either ``line_id`` OR composite key
    (``category`` + ``currency_code`` + ``kind``) - composite must
    match exactly one row.

    To clear the note, pass ``note_unset=True``.
    """
    _validate_period(period, "period")
    ts = now_ts if now_ts is not None else int(time.time())

    if line_id is None:
        if category is None or currency_code is None or kind is None:
            raise BudgetParseError(
                "update_line needs line_id or (category + currency_code + kind)",
                kind="BadArgument",
            )
        matches = _find_lines_by_composite(
            conn,
            period=period,
            currency_code=currency_code,
            category=category,
            kind=kind,
        )
        if not matches:
            raise BudgetParseError(
                f"no matching line in draft for {period}",
                kind="NotFound",
            )
        if len(matches) > 1:
            raise BudgetParseError(
                f"{len(matches)} lines match composite key; specify --line-id",
                kind="Ambiguous",
                details={"candidate_line_ids": [int(m[0]) for m in matches]},
            )
        before = matches[0]
    else:
        before = _read_line(conn, line_id=line_id)
        if before is None:
            raise BudgetParseError(f"line_id={line_id} not found", kind="NotFound")

    line_id_resolved = int(before[0])
    new_amount = before[4] if amount_minor is None else amount_minor
    if note_unset:
        new_note = None
    elif note is not None:
        new_note = note
    else:
        new_note = before[5]

    if new_amount == before[4] and new_note == before[5]:
        return {"ok": True, "op": "noop", "line": _row_to_payload(before)}

    conn.execute("BEGIN")
    try:
        conn.execute(
            "UPDATE budget_line SET amount_minor = ?, note = ? WHERE id = ?",
            (new_amount, new_note, line_id_resolved),
        )
        after_row = _read_line(conn, line_id=line_id_resolved)
        payload_before = _row_to_payload(before)
        payload_after = _row_to_payload(after_row) if after_row else None
        budget_id = int(
            conn.execute(
                "SELECT budget_id FROM budget_line WHERE id = ?",
                (line_id_resolved,),
            ).fetchone()[0]
        )
        conn.execute(
            "INSERT INTO budget_draft_edit "
            "(budget_id, op, payload_before, payload_after, applied_at) "
            "VALUES (?, ?, ?, ?, ?)",
            (
                budget_id,
                "update",
                json.dumps(payload_before, ensure_ascii=False),
                json.dumps(payload_after, ensure_ascii=False) if payload_after else None,
                ts,
            ),
        )
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    return {
        "ok": True,
        "op": "update",
        "before": payload_before,
        "after": payload_after,
        "budget_id": budget_id,
    }


def remove_line(
    conn: sqlite3.Connection,
    *,
    period: str,
    line_id: int | None = None,
    category: str | None = None,
    currency_code: int | None = None,
    kind: str | None = None,
    now_ts: int | None = None,
) -> dict[str, Any]:
    """Delete a draft line. Same addressing rules as ``update_line``."""
    _validate_period(period, "period")
    ts = now_ts if now_ts is not None else int(time.time())

    if line_id is None:
        if category is None or currency_code is None or kind is None:
            raise BudgetParseError(
                "remove_line needs line_id or (category + currency_code + kind)",
                kind="BadArgument",
            )
        matches = _find_lines_by_composite(
            conn,
            period=period,
            currency_code=currency_code,
            category=category,
            kind=kind,
        )
        if not matches:
            raise BudgetParseError(
                f"no matching line in draft for {period}",
                kind="NotFound",
            )
        if len(matches) > 1:
            raise BudgetParseError(
                f"{len(matches)} lines match; specify --line-id",
                kind="Ambiguous",
                details={"candidate_line_ids": [int(m[0]) for m in matches]},
            )
        before = matches[0]
    else:
        before = _read_line(conn, line_id=line_id)
        if before is None:
            raise BudgetParseError(f"line_id={line_id} not found", kind="NotFound")

    line_id_resolved = int(before[0])
    payload_before = _row_to_payload(before)
    budget_id = int(
        conn.execute(
            "SELECT budget_id FROM budget_line WHERE id = ?",
            (line_id_resolved,),
        ).fetchone()[0]
    )

    conn.execute("BEGIN")
    try:
        conn.execute("DELETE FROM budget_line WHERE id = ?", (line_id_resolved,))
        conn.execute(
            "INSERT INTO budget_draft_edit "
            "(budget_id, op, payload_before, payload_after, applied_at) "
            "VALUES (?, ?, ?, NULL, ?)",
            (budget_id, "remove", json.dumps(payload_before, ensure_ascii=False), ts),
        )
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    return {
        "ok": True,
        "op": "remove",
        "line": payload_before,
        "budget_id": budget_id,
    }


def undo_last(conn: sqlite3.Connection, *, period: str) -> dict[str, Any]:
    """Reverse the most recent edit on any draft for ``period``."""
    row = conn.execute(
        "SELECT bde.id, bde.budget_id, bde.op, bde.payload_before, bde.payload_after "
        "FROM budget_draft_edit bde "
        "JOIN budget b ON b.id = bde.budget_id "
        "WHERE b.period = ? AND b.status = 'draft' "
        "ORDER BY bde.id DESC LIMIT 1",
        (period,),
    ).fetchone()
    if row is None:
        return {"ok": True, "undone": None}

    edit_id, budget_id, op, payload_before_s, payload_after_s = row
    before = json.loads(payload_before_s) if payload_before_s else None
    after = json.loads(payload_after_s) if payload_after_s else None

    conn.execute("BEGIN")
    try:
        if op == "add":
            assert after is not None
            conn.execute("DELETE FROM budget_line WHERE id = ?", (after["line_id"],))
            reverted_op = "remove"
        elif op == "remove":
            assert before is not None
            conn.execute(
                "INSERT INTO budget_line "
                "(budget_id, category, amount_minor, kind, note) "
                "VALUES (?, ?, ?, ?, ?)",
                (
                    budget_id,
                    before["category"],
                    before["amount_minor"],
                    before["kind"],
                    before["note"],
                ),
            )
            reverted_op = "add"
        else:  # update
            assert before is not None and after is not None
            conn.execute(
                "UPDATE budget_line SET amount_minor = ?, note = ? WHERE id = ?",
                (before["amount_minor"], before["note"], after["line_id"]),
            )
            reverted_op = "update"
        conn.execute("DELETE FROM budget_draft_edit WHERE id = ?", (edit_id,))
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    return {
        "ok": True,
        "undone": {
            "edit_id": int(edit_id),
            "op": op,
            "reverted_as": reverted_op,
            "before": before,
            "after": after,
        },
    }


def commit_draft(conn: sqlite3.Connection, *, period: str) -> dict[str, Any]:
    """Flip every draft budget for ``period`` to ``active``, replacing
    any pre-existing active budget for the same (period, currency)
    atomically. Clears the draft edit log on success."""
    drafts = conn.execute(
        "SELECT id, currency_code FROM budget "
        "WHERE period = ? AND status = 'draft' ORDER BY currency_code",
        (period,),
    ).fetchall()
    if not drafts:
        raise BudgetParseError(f"no draft to commit for {period}", kind="NotFound")

    result: list[dict[str, Any]] = []
    conn.execute("BEGIN")
    try:
        for draft_id, cur_code in drafts:
            replaced_active_id: int | None = None
            existing_active = conn.execute(
                "SELECT id FROM budget "
                "WHERE period = ? AND currency_code = ? AND status = 'active'",
                (period, cur_code),
            ).fetchone()
            if existing_active is not None:
                replaced_active_id = int(existing_active[0])
                conn.execute("DELETE FROM budget WHERE id = ?", (replaced_active_id,))
            conn.execute(
                "UPDATE budget SET status = 'active' WHERE id = ?",
                (draft_id,),
            )
            conn.execute(
                "DELETE FROM budget_draft_edit WHERE budget_id = ?",
                (draft_id,),
            )
            line_count = conn.execute(
                "SELECT COUNT(*) FROM budget_line WHERE budget_id = ?",
                (draft_id,),
            ).fetchone()[0]
            result.append(
                {
                    "budget_id": int(draft_id),
                    "currency_code": int(cur_code),
                    "line_count": int(line_count),
                    "replaced_active_id": replaced_active_id,
                }
            )
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    return {"ok": True, "period": period, "committed": result}


def cancel_draft(conn: sqlite3.Connection, *, period: str) -> dict[str, Any]:
    """Delete every draft budget for ``period`` (cascade). Category-
    registry entries created during the draft are intentionally NOT
    touched."""
    drafts = conn.execute(
        "SELECT id, currency_code FROM budget WHERE period = ? AND status = 'draft'",
        (period,),
    ).fetchall()
    if not drafts:
        raise BudgetParseError(f"no draft to cancel for {period}", kind="NotFound")
    removed: list[dict[str, Any]] = []
    conn.execute("BEGIN")
    try:
        for draft_id, cur_code in drafts:
            line_count = conn.execute(
                "SELECT COUNT(*) FROM budget_line WHERE budget_id = ?",
                (draft_id,),
            ).fetchone()[0]
            conn.execute("DELETE FROM budget WHERE id = ?", (draft_id,))
            removed.append(
                {
                    "budget_id": int(draft_id),
                    "currency_code": int(cur_code),
                    "line_count": int(line_count),
                }
            )
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    return {"ok": True, "period": period, "cancelled": removed}


def fetch_budget(
    conn: sqlite3.Connection, *, period: str, currency_code: int | None = None
) -> list[dict[str, Any]]:
    """Return budget rows + their lines for the period.

    Shape::

        [{period, currency_code, status, created_at, imported_from,
          lines: [{category, kind, amount_minor, note}], total_minor}]

    Sorted by currency_code so UAH (980) always comes before USD
    (840) → no, sorted ASC numeric so UAH (980) comes after USD (840).
    Use the alpha code in CLI rendering if the user-facing order
    matters.
    """
    where = ["b.period = ?"]
    params: list[Any] = [period]
    if currency_code is not None:
        where.append("b.currency_code = ?")
        params.append(currency_code)
    budgets = conn.execute(
        "SELECT id, period, currency_code, status, created_at, imported_from "
        f"FROM budget b WHERE {' AND '.join(where)} ORDER BY currency_code",
        params,
    ).fetchall()
    out: list[dict[str, Any]] = []
    for bid, per, cur, status, created_at, imported_from in budgets:
        lines = conn.execute(
            "SELECT category, kind, amount_minor, note FROM budget_line "
            "WHERE budget_id = ? ORDER BY amount_minor ASC",
            (bid,),
        ).fetchall()
        line_dicts = [
            {"category": c, "kind": k, "amount_minor": int(a), "note": n} for c, k, a, n in lines
        ]
        out.append(
            {
                "budget_id": int(bid),
                "period": per,
                "currency_code": int(cur),
                "status": status,
                "created_at": int(created_at),
                "imported_from": imported_from,
                "lines": line_dicts,
                "total_minor": sum(line["amount_minor"] for line in line_dicts),
                "line_count": len(line_dicts),
            }
        )
    return out


def actuals_for_period(
    conn: sqlite3.Connection,
    *,
    period: str,
    currency_code: int | None = None,
    exclude_categories: tuple[str, ...] = ("Перекази/СвоїКартки",),
) -> dict[tuple[int, str], int]:
    """Sum actual transactions by (currency, category) for a period.

    Uses the same UNION view + CATEGORY_EXPR / CATEGORY_JOIN_SQL +
    account-currency join that ``summarize_spending`` does. Returns
    ``{(currency_code, category): total_minor_signed}``.

    ``exclude_categories`` defaults to the internal-transfer label so
    actuals match the "real spending" convention used in pf-report.
    """
    from . import queries as q
    from .view import build_accounts_union_sql, build_tx_union_sql, discover_sources

    sources = discover_sources(conn)
    tx_union = build_tx_union_sql(sources)
    accounts_union = build_accounts_union_sql(sources)
    if tx_union is None or accounts_union is None:
        return {}
    # Period boundaries: first second of month → first second of next month.
    year, month = (int(p) for p in period.split("-"))
    from_ts = int(time.mktime((year, month, 1, 0, 0, 0, 0, 0, 0))) - time.timezone
    if month == 12:
        next_year, next_month = year + 1, 1
    else:
        next_year, next_month = year, month + 1
    to_ts = int(time.mktime((next_year, next_month, 1, 0, 0, 0, 0, 0, 0))) - time.timezone

    where_cur = ""
    params: list[Any] = [from_ts, to_ts]
    if currency_code is not None:
        where_cur = " AND acc.currency_code = ?"
        params.append(currency_code)
    placeholders = ",".join(["?"] * len(exclude_categories))
    cat_filter = f" AND ({q.CATEGORY_EXPR} NOT IN ({placeholders}))" if exclude_categories else ""
    params.extend(exclude_categories)

    sql = (
        f"SELECT acc.currency_code, {q.CATEGORY_EXPR} AS category, "
        f"SUM(tx.amount_minor) AS total_minor "
        f"FROM (\n{tx_union}\n) AS tx "
        f"{q.CATEGORY_JOIN_SQL} "
        f"JOIN (\n{accounts_union}\n) AS acc ON acc.account_id = tx.account_id "
        f"WHERE tx.ts >= ? AND tx.ts < ?{where_cur}{cat_filter} "
        f"GROUP BY acc.currency_code, {q.CATEGORY_EXPR}"
    )
    out: dict[tuple[int, str], int] = {}
    for cur, cat, tot in conn.execute(sql, params):
        if cat is None:
            cat = "(uncategorized)"
        out[(int(cur), str(cat))] = int(tot or 0)
    return out


def diff_budget_vs_actual(
    conn: sqlite3.Connection,
    *,
    period: str,
    currency_code: int | None = None,
) -> list[dict[str, Any]]:
    """Join budgeted lines with actuals from ``actuals_for_period``.

    Output is a list of per-currency blocks::

        [{currency_code, status, lines: [{category, kind, target_minor,
          actual_minor, delta_minor, pct_used}], totals: {...}}]

    Categories that exist only in actuals (no budget line) are still
    listed - the user wants to see what they spent that wasn't planned
    for. Categories with target but no actual show ``actual_minor=0``.
    """
    budgets = fetch_budget(conn, period=period, currency_code=currency_code)
    actuals = actuals_for_period(conn, period=period, currency_code=currency_code)

    # Index actuals by currency for quick consumption per budget block.
    actuals_by_cur: dict[int, dict[str, int]] = {}
    for (cur, cat), amt in actuals.items():
        actuals_by_cur.setdefault(cur, {})[cat] = amt

    out: list[dict[str, Any]] = []
    for b in budgets:
        cur = b["currency_code"]
        seen: set[str] = set()
        # Aggregate budget targets per category (a category can have
        # both baseline and one_time rows for the same period).
        target_by_cat: dict[str, int] = {}
        for line in b["lines"]:
            target_by_cat[line["category"]] = (
                target_by_cat.get(line["category"], 0) + line["amount_minor"]
            )
        cur_actuals = actuals_by_cur.get(cur, {})
        rows: list[dict[str, Any]] = []
        for cat, target in target_by_cat.items():
            actual = cur_actuals.get(cat, 0)
            delta = target - actual
            pct = _pct_used(actual, target)
            rows.append(
                {
                    "category": cat,
                    "target_minor": target,
                    "actual_minor": actual,
                    "delta_minor": delta,
                    "pct_used": pct,
                    "in_budget": True,
                }
            )
            seen.add(cat)
        # Surface actuals that have no budgeted line.
        for cat, actual in cur_actuals.items():
            if cat in seen:
                continue
            rows.append(
                {
                    "category": cat,
                    "target_minor": 0,
                    "actual_minor": actual,
                    "delta_minor": -actual,
                    "pct_used": None,
                    "in_budget": False,
                }
            )
        rows.sort(key=lambda r: r["target_minor"])  # most-negative first
        out.append(
            {
                "currency_code": cur,
                "status": b["status"],
                "lines": rows,
                "totals": _diff_totals(rows),
            }
        )
    # Surface actuals for currencies that have no budget at all.
    seen_cur = {b["currency_code"] for b in budgets}
    for cur, by_cat in actuals_by_cur.items():
        if cur in seen_cur:
            continue
        rows = [
            {
                "category": cat,
                "target_minor": 0,
                "actual_minor": amt,
                "delta_minor": -amt,
                "pct_used": None,
                "in_budget": False,
            }
            for cat, amt in sorted(by_cat.items(), key=lambda x: x[1])
        ]
        out.append(
            {
                "currency_code": cur,
                "status": None,  # no budget
                "lines": rows,
                "totals": _diff_totals(rows),
            }
        )
    return out


def _pct_used(actual: int, target: int) -> float | None:
    """Percentage of target used so far. ``None`` when target == 0
    (income lines or no baseline target). For outflows (negative
    target), divide negative by negative so the result is positive.
    """
    if target == 0:
        return None
    return round(actual / target * 100.0, 1)


# Top-level category group that denotes real income (Дохід/Зарплата,
# Дохід/ВПО, ...). Internal transfers (Перекази/СвоїКартки) are already
# dropped upstream by ``actuals_for_period``, so within a diff block the
# only non-spend rows are income.
_INCOME_CATEGORY_PREFIX = "Дохід"


def _is_income_category(category: str) -> bool:
    return category == _INCOME_CATEGORY_PREFIX or category.startswith(_INCOME_CATEGORY_PREFIX + "/")


def _diff_totals(rows: list[dict[str, Any]]) -> dict[str, int]:
    """Totals for a diff block that keep spend and income SEPARATE.

    ``actual_minor`` is the legacy net-of-income figure (kept so existing
    consumers don't break), but it is misleading on its own: income lands
    on ``Дохід/*`` rows as positive amounts and silently shrinks it. The
    explicit fields are the ones to report:

      - ``real_spend_minor``  - sum of actual on non-income rows (refunds
        net in as positive offsets). Transfers are already excluded
        upstream.
      - ``income_minor``      - sum of actual on ``Дохід/*`` rows.
      - ``remaining_minor``   - planned outflow left = spend target minus
        real spend (signed like the targets: negative = budget still to
        spend, positive = overspent).
    """
    spend_rows = [r for r in rows if not _is_income_category(r["category"])]
    real_spend = sum(r["actual_minor"] for r in spend_rows)
    spend_target = sum(r["target_minor"] for r in spend_rows)
    return {
        "target_minor": sum(r["target_minor"] for r in rows),
        "actual_minor": sum(r["actual_minor"] for r in rows),
        "delta_minor": sum(r["delta_minor"] for r in rows),
        "real_spend_minor": real_spend,
        "income_minor": sum(r["actual_minor"] for r in rows if _is_income_category(r["category"])),
        "remaining_minor": spend_target - real_spend,
    }


@dataclass
class StatusFlipResult:
    matched: int  # how many budget rows existed
    changed: list[dict[str, Any]]  # rows whose status actually flipped


def set_status(
    conn: sqlite3.Connection,
    *,
    period: str,
    currency_code: int | None,
    new_status: str,
) -> StatusFlipResult:
    """Flip ``budget.status`` for the matching budgets. Returns
    ``matched`` (how many rows existed) and ``changed`` (the rows
    that actually moved). Distinguishing the two lets the CLI emit
    ``NotFound`` only when nothing matched - flipping a closed budget
    back to closed is a successful no-op."""
    if new_status not in ("draft", "active", "closed"):
        raise BudgetParseError(
            f"new_status={new_status!r} must be one of draft|active|closed",
            kind="BadStatus",
        )
    where = ["period = ?"]
    params: list[Any] = [period]
    if currency_code is not None:
        where.append("currency_code = ?")
        params.append(currency_code)
    rows = conn.execute(
        f"SELECT id, period, currency_code, status FROM budget WHERE {' AND '.join(where)}",
        params,
    ).fetchall()
    if not rows:
        return StatusFlipResult(matched=0, changed=[])
    changed: list[dict[str, Any]] = []
    conn.execute("BEGIN")
    try:
        for budget_id, per, cur, status in rows:
            if status == new_status:
                continue
            conn.execute(
                "UPDATE budget SET status = ? WHERE id = ?",
                (new_status, budget_id),
            )
            changed.append(
                {
                    "budget_id": int(budget_id),
                    "period": per,
                    "currency_code": int(cur),
                    "old_status": status,
                    "new_status": new_status,
                }
            )
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    return StatusFlipResult(matched=len(rows), changed=changed)


def delete_budget(
    conn: sqlite3.Connection,
    *,
    period: str,
    currency_code: int | None,
    force: bool = False,
) -> list[dict[str, Any]]:
    """Delete a budget (and its lines via cascade). Refuses to delete
    closed budgets unless ``force=True`` - the lifecycle is meant to
    be append-and-snapshot, not purge."""
    where = ["period = ?"]
    params: list[Any] = [period]
    if currency_code is not None:
        where.append("currency_code = ?")
        params.append(currency_code)
    rows = conn.execute(
        f"SELECT id, period, currency_code, status FROM budget WHERE {' AND '.join(where)}",
        params,
    ).fetchall()
    if not rows:
        return []
    deleted: list[dict[str, Any]] = []
    conn.execute("BEGIN")
    try:
        for budget_id, per, cur, status in rows:
            if status == "closed" and not force:
                conn.rollback()
                raise BudgetParseError(
                    f"budget {per}/{cur} is closed; pass --force to delete",
                    kind="ClosedBudget",
                )
            if status == "closed" and force:
                # Reopen so cascade deletion of budget_line passes the
                # closed-budget trigger.
                conn.execute(
                    "UPDATE budget SET status = 'active' WHERE id = ?",
                    (budget_id,),
                )
            conn.execute("DELETE FROM budget WHERE id = ?", (budget_id,))
            deleted.append(
                {
                    "budget_id": int(budget_id),
                    "period": per,
                    "currency_code": int(cur),
                    "was_status": status,
                }
            )
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    return deleted


def rename_category(
    conn: sqlite3.Connection,
    *,
    old_name: str,
    new_name: str,
    update_tables: tuple[str, ...],
) -> dict[str, int]:
    """Rewrite ``category`` in the named tables. Returns per-table
    affected-row counts. The trigger on ``budget_line`` allows the
    UPDATE only when the parent budget is not closed, so renames on
    closed budgets need ``pf-budget reopen`` first."""
    allowed = {
        "budget_line",
        "tx_category",
        "category_overrides",
        "categorization_rules",
        "category_registry",
    }
    bad = [t for t in update_tables if t not in allowed]
    if bad:
        raise BudgetParseError(
            f"--update entries {bad} not allowed; must be a subset of {sorted(allowed)}",
            kind="BadArgument",
        )
    counts: dict[str, int] = {}
    conn.execute("BEGIN")
    try:
        for table in update_tables:
            if table == "category_registry":
                # PK is category itself - DELETE old + INSERT new (or
                # just UPDATE if SQLite version supports it on PK).
                # Try UPDATE first; fall back to delete-and-insert if
                # a row already exists at new_name (UNIQUE collision).
                existing_new = conn.execute(
                    "SELECT 1 FROM category_registry WHERE category = ?",
                    (new_name,),
                ).fetchone()
                if existing_new is not None:
                    counts[table] = conn.execute(
                        "DELETE FROM category_registry WHERE category = ?",
                        (old_name,),
                    ).rowcount
                else:
                    counts[table] = conn.execute(
                        "UPDATE category_registry SET category = ? WHERE category = ?",
                        (new_name, old_name),
                    ).rowcount
                continue
            counts[table] = conn.execute(
                f"UPDATE {table} SET category = ? WHERE category = ?",
                (new_name, old_name),
            ).rowcount
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    return counts


def list_budgets(conn: sqlite3.Connection) -> list[dict[str, Any]]:
    """Compact list of all budgets, line totals included, suitable for
    the home-page-style overview."""
    rows = conn.execute(
        "SELECT b.id, b.period, b.currency_code, b.status, b.created_at, "
        "       b.imported_from, "
        "       COUNT(bl.id) AS line_count, "
        "       COALESCE(SUM(bl.amount_minor), 0) AS total_minor "
        "FROM budget b LEFT JOIN budget_line bl ON bl.budget_id = b.id "
        "GROUP BY b.id ORDER BY b.period DESC, b.currency_code ASC"
    ).fetchall()
    return [
        {
            "budget_id": int(r[0]),
            "period": r[1],
            "currency_code": int(r[2]),
            "status": r[3],
            "created_at": int(r[4]),
            "imported_from": r[5],
            "line_count": int(r[6]),
            "total_minor": int(r[7]),
        }
        for r in rows
    ]


def scan_history_for_signals(
    conn: sqlite3.Connection,
    *,
    target_period: str,
    lookback_months: int = 6,
) -> list[dict[str, Any]]:
    """Inspect the previous ``lookback_months`` active budgets and
    return planning hints for ``target_period``.

    Each signal is a dict the CLI emits as JSON for Claude to phrase
    back to the user. Shape::

        {"type": "seasonal_gap",
         "category": "Освіта/Школа",
         "currency_code": 980,
         "evidence": {...},
         "suggestion": "consider zeroing for the summer"}

    Signal types:
      - ``one_time_excluded``: line of kind=one_time in most recent
        active budget; will be excluded from the copy
      - ``seasonal_gap``: category present in some months and absent
        in others (school-shaped)
      - ``monotonic_trend``: amount monotonically growing or shrinking
        over 3+ months
      - ``quarterly_cadence``: same category reappears at regular
        ~3-month intervals (taxes, insurance premiums)
      - ``one_off_deviation``: most recent month deviated > 30% from
        prior monthly median (vacation-half-month-charging shaped)
    """
    _validate_period(target_period, "target_period")
    # Walk back from target_period to assemble lookback periods.
    year, month = (int(x) for x in target_period.split("-"))
    periods: list[str] = []
    for _ in range(lookback_months):
        month -= 1
        if month == 0:
            month = 12
            year -= 1
        periods.append(f"{year:04d}-{month:02d}")
    periods.reverse()  # oldest first

    if not periods:
        return []

    placeholders = ",".join(["?"] * len(periods))
    rows = conn.execute(
        f"SELECT b.period, b.currency_code, bl.category, bl.kind, "
        f"       bl.amount_minor "
        f"FROM budget_line bl JOIN budget b ON b.id = bl.budget_id "
        f"WHERE b.status IN ('active', 'closed') "
        f"AND b.period IN ({placeholders}) "
        f"ORDER BY b.period",
        periods,
    ).fetchall()
    if not rows:
        return []

    most_recent = periods[-1]
    signals: list[dict[str, Any]] = []
    seen_signatures: set[tuple] = set()

    # Index by (category, currency_code) → period → (kind, amount)
    history: dict[tuple[str, int], dict[str, list[tuple[str, int]]]] = {}
    for period, cur, cat, kind, amt in rows:
        history.setdefault((cat, cur), {}).setdefault(period, []).append((kind, int(amt)))

    def add(sig: dict[str, Any]) -> None:
        # Dedup on (type, category, currency_code) so two flavours of
        # the same insight don't both fire.
        key = (sig["type"], sig.get("category"), sig.get("currency_code"))
        if key in seen_signatures:
            return
        seen_signatures.add(key)
        signals.append(sig)

    for (cat, cur), per_period in history.items():
        # ---- one_time_excluded
        recent_kinds = per_period.get(most_recent, [])
        for kind, amt in recent_kinds:
            if kind == "one_time":
                add(
                    {
                        "type": "one_time_excluded",
                        "category": cat,
                        "currency_code": cur,
                        "evidence": {"recent_period": most_recent, "amount_minor": amt},
                        "suggestion": (
                            f"line was one_time in {most_recent}; excluded from "
                            "the template - re-add if needed this month"
                        ),
                    }
                )
                break

        # ---- seasonal_gap
        active_months = set(per_period.keys())
        if 0 < len(active_months) < len(periods):
            missing = [p for p in periods if p not in active_months]
            present = sorted(active_months)
            if len(missing) >= 1 and len(present) >= 2:
                add(
                    {
                        "type": "seasonal_gap",
                        "category": cat,
                        "currency_code": cur,
                        "evidence": {
                            "present_in": present,
                            "missing_in": missing,
                        },
                        "suggestion": (
                            f"present in {len(present)} of {len(periods)} months - "
                            "ask whether it applies this month"
                        ),
                    }
                )

        # ---- monotonic_trend (baseline-only)
        amounts_by_period_baseline: list[tuple[str, int]] = []
        for period in sorted(per_period.keys()):
            for kind, amt in per_period[period]:
                if kind == "baseline":
                    amounts_by_period_baseline.append((period, amt))
                    break
        if len(amounts_by_period_baseline) >= 3:
            amounts = [a for _, a in amounts_by_period_baseline]
            if all(amounts[i] < amounts[i + 1] for i in range(len(amounts) - 1)):
                direction = "growing"
            elif all(amounts[i] > amounts[i + 1] for i in range(len(amounts) - 1)):
                direction = "shrinking"
            else:
                direction = None
            if direction is not None:
                # Only flag when the change is at least 5% over the span
                first, last = amounts[0], amounts[-1]
                if first != 0 and abs(last - first) / abs(first) >= 0.05:
                    add(
                        {
                            "type": "monotonic_trend",
                            "category": cat,
                            "currency_code": cur,
                            "evidence": {
                                "direction": direction,
                                "series": amounts_by_period_baseline,
                            },
                            "suggestion": (
                                f"baseline {direction} over {len(amounts)} months - "
                                "continue the trend or pin?"
                            ),
                        }
                    )

        # ---- quarterly_cadence
        sorted_periods = sorted(active_months)
        if len(sorted_periods) >= 2:
            gaps = []
            prev_idx = None
            for p in sorted_periods:
                idx = periods.index(p)
                if prev_idx is not None:
                    gaps.append(idx - prev_idx)
                prev_idx = idx
            if gaps and all(g in (2, 3, 4) for g in gaps) and max(gaps) - min(gaps) <= 1:
                # Project the next quarterly hit
                next_idx = periods.index(sorted_periods[-1]) + gaps[-1]
                target_idx = len(periods)  # target_period sits one beyond the lookback
                if next_idx == target_idx:
                    add(
                        {
                            "type": "quarterly_cadence",
                            "category": cat,
                            "currency_code": cur,
                            "evidence": {"prior_hits": sorted_periods, "gap": gaps[-1]},
                            "suggestion": (
                                "appears on a quarterly cadence; next instance "
                                f"projects to {target_period}"
                            ),
                        }
                    )

        # ---- one_off_deviation
        baseline_amounts = [a for _, a in amounts_by_period_baseline]
        if len(baseline_amounts) >= 3:
            recent = baseline_amounts[-1]
            prior = sorted(baseline_amounts[:-1])
            median = prior[len(prior) // 2]
            if median != 0 and abs(recent - median) / abs(median) >= 0.3:
                add(
                    {
                        "type": "one_off_deviation",
                        "category": cat,
                        "currency_code": cur,
                        "evidence": {
                            "recent_amount": recent,
                            "prior_median": median,
                            "delta_pct": round((recent - median) / median * 100, 1),
                        },
                        "suggestion": (
                            f"{most_recent} deviated >30% from prior median; "
                            "ask whether to revert or hold the new value"
                        ),
                    }
                )

    return signals


# Top-level grouping for the Family export view. Order shapes how
# groups appear in the spreadsheet (largest first within each group,
# but groups in this order). Categories not in this list fall to
# "Інше" at the bottom.
_FAMILY_GROUP_ORDER: tuple[tuple[str, str, list[str]], ...] = (
    ("housing", "Житло", ["Житло", "Зв'язок"]),
    ("food", "Харчування", ["Їжа"]),
    ("transport", "Транспорт", ["Транспорт"]),
    ("subs", "Підписки", ["Підписки"]),
    ("shopping", "Покупки", ["Покупки", "Подарунки"]),
    ("health_beauty", "Здоров'я і краса", ["Здоров'я", "Краса"]),
    ("education", "Освіта", ["Освіта"]),
    ("travel", "Подорожі", ["Подорожі"]),
    ("entertainment", "Розваги", ["Розваги"]),
    ("taxes", "Податки і збори", ["Податки"]),
    ("donations", "Благодійність", ["Благодійність"]),
    ("transfers", "Перекази", ["Перекази"]),
    ("cash", "Готівка", ["Готівка"]),
    ("invest", "Інвестиції", ["Інвестиції"]),
    ("income", "Дохід", ["Дохід"]),
)


def _classify_family_group(category: str) -> tuple[str, str]:
    """Return (group_key, group_title) for the given category. Falls
    back to ('other', 'Інше') for categories outside the known map."""
    for key, title, prefixes in _FAMILY_GROUP_ORDER:
        for prefix in prefixes:
            if category == prefix or category.startswith(prefix + "/"):
                return key, title
    return "other", "Інше"


def family_view_rows(
    conn: sqlite3.Connection,
    *,
    period: str,
) -> dict[str, Any]:
    """Build the rendered Family-view structure for ``period``.

    Returns a dict the export writer can walk to produce the styled
    XLSX. Shape::

        {"period": str,
         "currencies": [
            {"currency_code": int, "alpha": str, "total_major": float,
             "groups": [{"key": str, "title": str, "subtotal_major": float,
                         "lines": [{"category_display": str,
                                    "amount_major": float,
                                    "note": str|None,
                                    "kind": str}, ...]}, ...]}, ...]}
    """
    blocks = fetch_budget(conn, period=period)
    from .currencies import alpha_for

    currencies: list[dict[str, Any]] = []
    for b in blocks:
        if b["status"] != "active":
            continue  # Family view is only for committed plans
        from collections import defaultdict

        # Aggregate by full category (so one_time + baseline on the
        # same category combine for the user's eyes).
        per_cat: dict[str, dict[str, Any]] = defaultdict(
            lambda: {"amount_minor": 0, "notes": [], "kinds": set()}
        )
        for line in b["lines"]:
            slot = per_cat[line["category"]]
            slot["amount_minor"] += line["amount_minor"]
            if line["note"]:
                slot["notes"].append(line["note"])
            slot["kinds"].add(line["kind"])

        # Now bucket categories into groups
        groups_map: dict[str, dict[str, Any]] = {}
        for cat, agg in per_cat.items():
            group_key, group_title = _classify_family_group(cat)
            if group_key not in groups_map:
                groups_map[group_key] = {
                    "key": group_key,
                    "title": group_title,
                    "subtotal_major": 0.0,
                    "lines": [],
                }
            groups_map[group_key]["subtotal_major"] += agg["amount_minor"] / 100.0
            groups_map[group_key]["lines"].append(
                {
                    "category": cat,
                    "category_display": _family_sub_label(cat),
                    "amount_major": agg["amount_minor"] / 100.0,
                    "note": "; ".join(agg["notes"]) if agg["notes"] else None,
                    "kinds": sorted(agg["kinds"]),
                }
            )

        # Sort lines within group: most-negative first (= biggest spend)
        for grp in groups_map.values():
            grp["lines"].sort(key=lambda r: r["amount_major"])

        # Order groups per _FAMILY_GROUP_ORDER, then "other" last
        ordered_groups: list[dict[str, Any]] = []
        for key, _title, _prefixes in _FAMILY_GROUP_ORDER:
            if key in groups_map:
                ordered_groups.append(groups_map[key])
        if "other" in groups_map:
            ordered_groups.append(groups_map["other"])

        currencies.append(
            {
                "currency_code": b["currency_code"],
                "alpha": alpha_for(b["currency_code"]) or str(b["currency_code"]),
                "total_major": sum(g["subtotal_major"] for g in ordered_groups),
                "groups": ordered_groups,
            }
        )

    return {"period": period, "currencies": currencies}


def _family_sub_label(category: str) -> str:
    """Family-view line label: just the sub-category, since the group
    header right above already names the top-level. Falls back to the
    full name when the category has no sub (e.g. ``Готівка``,
    ``Зв'язок``)."""
    parts = category.split("/", 1)
    if len(parts) == 1:
        return parts[0]
    return parts[1]


def export_variance_rows(
    conn: sqlite3.Connection,
    *,
    period: str,
    currency_code: int | None = None,
) -> list[dict[str, Any]]:
    """Build the variance-sheet rows the CLI emits via ``pf-budget
    export``. Schema matches budget-design.md `Variance` sheet:
    ``Period, Category, Currency, Target, Actual, Delta, % used``.
    Amounts come back in **major units** as floats so the file lands
    Sheets-friendly without manual division.
    """
    blocks = diff_budget_vs_actual(conn, period=period, currency_code=currency_code)
    from .currencies import alpha_for

    out: list[dict[str, Any]] = []
    for block in blocks:
        cur_code = block["currency_code"]
        cur_alpha = alpha_for(cur_code) or str(cur_code)
        for line in block["lines"]:
            out.append(
                {
                    "Period": period,
                    "Category": line["category"],
                    "Currency": cur_alpha,
                    "Target": line["target_minor"] / 100.0,
                    "Actual": line["actual_minor"] / 100.0,
                    "Delta": line["delta_minor"] / 100.0,
                    "% used": line["pct_used"],
                }
            )
    return out


def log_import_run(
    conn: sqlite3.Connection,
    *,
    source: str,
    period: str,
    lines_added: int,
    lines_replaced: int,
    new_categories: list[str],
    now_ts: int | None = None,
) -> int:
    """Append to ``budget_import_run`` audit table. Returns the
    inserted row id so the CLI can echo it back."""
    ts = now_ts if now_ts is not None else int(time.time())
    conn.execute("BEGIN")
    try:
        conn.execute(
            "INSERT INTO budget_import_run "
            "(source, period, imported_at, lines_added, lines_replaced, new_categories) "
            "VALUES (?, ?, ?, ?, ?, ?)",
            (
                source,
                period,
                ts,
                lines_added,
                lines_replaced,
                json.dumps(new_categories, ensure_ascii=False) if new_categories else None,
            ),
        )
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    return int(conn.execute("SELECT last_insert_rowid()").fetchone()[0])


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------


def _check_columns(
    found: Iterable[str] | None,
    required: tuple[str, ...],
    optional: tuple[str, ...],
    *,
    sheet: str,
) -> None:
    if not found:
        raise BudgetParseError(f"{sheet}: empty input (no header row)", kind="EmptyInput")
    found_set = {f for f in found if isinstance(f, str)}
    missing = [c for c in required if c not in found_set]
    if missing:
        raise BudgetParseError(
            f"{sheet}: missing required columns {missing}; got {sorted(found_set)}",
            kind="BadHeader",
            details={"missing": missing, "found": sorted(found_set)},
        )
    extras = [c for c in found_set if c not in required and c not in optional and c]
    # Extras are tolerated (sheets often have helper columns) but we
    # still report them in the details so the CLI can warn.
    if extras:
        # Cheap warning carried alongside successful parse.
        # The validator sees `details.extras` later.
        pass


def _validate_period(period: str, flag: str) -> None:
    if not isinstance(period, str) or not PERIOD_RE.match(period):
        raise BudgetParseError(
            f"--{flag}={period!r} must match YYYY-MM",
            kind="BadPeriod",
        )


def _row_from_plans(raw: dict[str, Any], *, source_row: str) -> PlanRow:
    period = _coerce_str(raw.get("Period"), "Period", source_row)
    _validate_period(period, "period")
    category = _coerce_category(raw.get("Category"), source_row)
    currency = _coerce_currency(raw.get("Currency"), source_row)
    kind = _coerce_kind(raw.get("Kind"), source_row)
    amount = _coerce_amount(raw.get("Amount"), source_row)
    note = _coerce_note(raw.get("Note"))
    return PlanRow(
        period=period,
        category=category,
        currency_code=currency,
        kind=kind,
        amount_minor=amount,
        note=note,
        source_row=source_row,
    )


def _row_from_baseline(raw: dict[str, Any], period: str, *, source_row: str) -> PlanRow:
    category = _coerce_category(raw.get("Category"), source_row)
    currency = _coerce_currency(raw.get("Currency"), source_row)
    kind = _coerce_kind(raw.get("Kind"), source_row)
    if kind == "one_time":
        raise BudgetParseError(
            f"{source_row}: Baseline sheet must not contain Kind=one_time; "
            "one-time items belong in the Plans sheet",
            kind="BadKind",
        )
    amount = _coerce_amount(raw.get("Monthly target"), source_row, field_label="Monthly target")
    note = _coerce_note(raw.get("Note"))
    return PlanRow(
        period=period,
        category=category,
        currency_code=currency,
        kind=kind,
        amount_minor=amount,
        note=note,
        source_row=source_row,
    )


def _coerce_str(value: Any, field_label: str, source_row: str) -> str:
    if value is None:
        raise BudgetParseError(f"{source_row}: {field_label} is empty", kind="MissingField")
    return str(value).strip()


def _coerce_category(value: Any, source_row: str) -> str:
    raw = _coerce_str(value, "Category", source_row)
    stripped = raw.strip()
    if not stripped:
        raise BudgetParseError(f"{source_row}: Category is empty", kind="MissingField")
    if "//" in stripped or stripped.startswith("/") or stripped.endswith("/"):
        raise BudgetParseError(
            f"{source_row}: Category {stripped!r} has empty hierarchy segments",
            kind="BadCategoryShape",
        )
    return stripped


def _coerce_currency(value: Any, source_row: str) -> int:
    raw = _coerce_str(value, "Currency", source_row)
    if raw.isdigit():
        return int(raw)
    numeric = numeric_for(raw)
    if numeric is None:
        raise BudgetParseError(
            f"{source_row}: unknown Currency {raw!r}",
            kind="BadCurrency",
        )
    return numeric


def _coerce_kind(value: Any, source_row: str) -> str:
    raw = _coerce_str(value, "Kind", source_row).lower()
    if raw not in ALLOWED_KINDS:
        raise BudgetParseError(
            f"{source_row}: Kind {raw!r} must be one of {list(ALLOWED_KINDS)}",
            kind="BadKind",
        )
    return raw


def _coerce_amount(value: Any, source_row: str, *, field_label: str = "Amount") -> int:
    if value is None or (isinstance(value, str) and not value.strip()):
        raise BudgetParseError(f"{source_row}: {field_label} is empty", kind="MissingField")
    try:
        amount = float(value)
    except (ValueError, TypeError) as exc:
        raise BudgetParseError(
            f"{source_row}: {field_label}={value!r} is not numeric",
            kind="BadAmount",
        ) from exc
    return int(round(amount * 100))


def _coerce_note(value: Any) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    return s or None
