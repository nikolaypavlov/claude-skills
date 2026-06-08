"""History scanner signals + Family export view."""

from __future__ import annotations

import json
from pathlib import Path

import pytest

from pf_skill.budget_cli import main
from pf_skill.common import budget as bud
from pf_skill.common.store import open_db


def _run(argv: list[str], capsys: pytest.CaptureFixture[str]) -> tuple[int, dict, dict]:
    rc = main(argv)
    captured = capsys.readouterr()
    out: dict = json.loads(captured.out) if captured.out.strip() else {}
    err: dict = {}
    if captured.err.strip():
        try:
            err = json.loads(captured.err.splitlines()[-1])
        except json.JSONDecodeError:
            err = {"raw": captured.err}
    return rc, out, err


def _seed(db: Path, period: str, *rows: tuple) -> None:
    """rows: (category, currency_code, kind, amount_minor[, note])."""
    conn = open_db(db)
    plan = [
        bud.PlanRow(
            period,
            r[0],
            r[1],
            r[2],
            r[3],
            r[4] if len(r) > 4 else None,
        )
        for r in rows
    ]
    bud.materialise_budget(conn, period=period, rows=plan, source="seed")
    conn.close()


# --- scanner ---------------------------------------------------------------


def test_signal_one_time_excluded(empty_db: Path) -> None:
    _seed(empty_db, "2026-06", ("Подорожі/Готелі", 980, "one_time", -2728000))
    conn = open_db(empty_db)
    signals = bud.scan_history_for_signals(conn, target_period="2026-07")
    conn.close()
    one_time = [s for s in signals if s["type"] == "one_time_excluded"]
    assert len(one_time) == 1
    assert one_time[0]["category"] == "Подорожі/Готелі"


def test_signal_seasonal_gap(empty_db: Path) -> None:
    _seed(empty_db, "2026-04", ("Освіта/Школа", 980, "baseline", -1560000))
    _seed(empty_db, "2026-05", ("Освіта/Школа", 980, "baseline", -1560000))
    _seed(empty_db, "2026-06", ("Освіта/Школа", 980, "baseline", -1560000))
    # 2026-03 has no school
    conn = open_db(empty_db)
    signals = bud.scan_history_for_signals(conn, target_period="2026-07")
    conn.close()
    seasonal = [s for s in signals if s["type"] == "seasonal_gap"]
    assert any(s["category"] == "Освіта/Школа" for s in seasonal)


def test_signal_monotonic_trend_below_threshold(empty_db: Path) -> None:
    """Small drifts (<5% over the lookback) must not fire - the user
    cares about meaningful trends, not noise."""
    _seed(empty_db, "2026-04", ("Підписки/AI", 980, "baseline", -519500))
    _seed(empty_db, "2026-05", ("Підписки/AI", 980, "baseline", -526300))
    _seed(empty_db, "2026-06", ("Підписки/AI", 980, "baseline", -533300))
    conn = open_db(empty_db)
    signals = bud.scan_history_for_signals(conn, target_period="2026-07")
    conn.close()
    trends = [
        s for s in signals
        if s["type"] == "monotonic_trend" and s["category"] == "Підписки/AI"
    ]
    assert trends == []


def test_signal_monotonic_trend_large_enough(empty_db: Path) -> None:
    _seed(empty_db, "2026-04", ("Розваги/X", 980, "baseline", -100000))
    _seed(empty_db, "2026-05", ("Розваги/X", 980, "baseline", -150000))
    _seed(empty_db, "2026-06", ("Розваги/X", 980, "baseline", -200000))
    conn = open_db(empty_db)
    signals = bud.scan_history_for_signals(conn, target_period="2026-07")
    conn.close()
    trends = [s for s in signals if s["type"] == "monotonic_trend"]
    assert any(s["category"] == "Розваги/X" for s in trends)
    matching = next(s for s in trends if s["category"] == "Розваги/X")
    assert matching["evidence"]["direction"] in ("growing", "shrinking")


def test_signal_one_off_deviation(empty_db: Path) -> None:
    _seed(empty_db, "2026-04", ("Транспорт/Зарядка", 980, "baseline", -300000))
    _seed(empty_db, "2026-05", ("Транспорт/Зарядка", 980, "baseline", -300000))
    _seed(empty_db, "2026-06", ("Транспорт/Зарядка", 980, "baseline", -150000))  # half month
    conn = open_db(empty_db)
    signals = bud.scan_history_for_signals(conn, target_period="2026-07")
    conn.close()
    devs = [s for s in signals if s["type"] == "one_off_deviation"]
    assert any(s["category"] == "Транспорт/Зарядка" for s in devs)


def test_cli_plan_suggest_returns_signals(
    empty_db: Path, capsys: pytest.CaptureFixture[str]
) -> None:
    _seed(empty_db, "2026-06", ("Подорожі/Готелі", 980, "one_time", -10000))
    rc, out, err = _run(
        ["plan", "suggest", "--period", "2026-07", "--db", str(empty_db)],
        capsys,
    )
    assert rc == 0, err
    assert any(s["type"] == "one_time_excluded" for s in out["signals"])


# --- Family view ------------------------------------------------------------


def test_family_view_groups_categories(empty_db: Path) -> None:
    _seed(
        empty_db,
        "2026-07",
        ("Їжа/Ресторани", 980, "baseline", -900000),
        ("Їжа/Продукти", 980, "baseline", -800000),
        ("Житло/Оренда", 980, "baseline", -1950000),
        ("Підписки/AI", 980, "baseline", -530000),
        ("Інвестиції/Облігації", 840, "baseline", -170000),
    )
    conn = open_db(empty_db)
    data = bud.family_view_rows(conn, period="2026-07")
    conn.close()

    assert data["period"] == "2026-07"
    # Two currencies
    by_cur = {c["currency_code"]: c for c in data["currencies"]}
    assert 980 in by_cur and 840 in by_cur

    uah = by_cur[980]
    group_titles = [g["title"] for g in uah["groups"]]
    # Housing should come before Food per _FAMILY_GROUP_ORDER
    assert "Житло" in group_titles
    assert "Харчування" in group_titles
    assert group_titles.index("Житло") < group_titles.index("Харчування")

    # Food group has both ресторани and продукти
    food = next(g for g in uah["groups"] if g["title"] == "Харчування")
    cats = {line["category"] for line in food["lines"]}
    assert cats == {"Їжа/Ресторани", "Їжа/Продукти"}


def test_family_view_line_labels_are_sub_only(empty_db: Path) -> None:
    """Line labels render as just the sub-category - the group header
    already states the top-level. Categories without a sub fall back
    to the full name."""
    _seed(
        empty_db,
        "2026-07",
        ("Їжа/Ресторани", 980, "baseline", -900000),
        ("Готівка", 980, "baseline", -250000),  # no sub
    )
    conn = open_db(empty_db)
    data = bud.family_view_rows(conn, period="2026-07")
    conn.close()
    labels = {
        line["category"]: line["category_display"]
        for grp in data["currencies"][0]["groups"]
        for line in grp["lines"]
    }
    assert labels["Їжа/Ресторани"] == "Ресторани"
    assert labels["Готівка"] == "Готівка"


def test_family_export_writes_xlsx_with_two_sheets(
    empty_db: Path, tmp_path: Path, capsys: pytest.CaptureFixture[str]
) -> None:
    pytest.importorskip("openpyxl")
    from openpyxl import load_workbook

    _seed(
        empty_db,
        "2026-07",
        ("Їжа/Ресторани", 980, "baseline", -900000),
        ("Інвестиції/Облігації", 840, "baseline", -170000),
    )
    out_path = tmp_path / "family.xlsx"
    rc, out, err = _run(
        [
            "export",
            "--period",
            "2026-07",
            "--view",
            "family",
            "--out",
            str(out_path),
            "--db",
            str(empty_db),
        ],
        capsys,
    )
    assert rc == 0, err
    assert out["view"] == "family"
    wb = load_workbook(out_path)
    assert "Огляд" in wb.sheetnames
    assert "Деталі" in wb.sheetnames


def test_family_export_csv_rejected(
    empty_db: Path, tmp_path: Path, capsys: pytest.CaptureFixture[str]
) -> None:
    _seed(empty_db, "2026-07", ("X", 980, "baseline", -100))
    rc, _, err = _run(
        [
            "export",
            "--period",
            "2026-07",
            "--view",
            "family",
            "--out",
            str(tmp_path / "x.csv"),
            "--format",
            "csv",
            "--db",
            str(empty_db),
        ],
        capsys,
    )
    assert rc == 1
    assert "xlsx" in err["error"]


def test_export_plan_csv(
    empty_db: Path, tmp_path: Path, capsys: pytest.CaptureFixture[str]
) -> None:
    _seed(empty_db, "2026-07", ("Їжа/Ресторани", 980, "baseline", -900000))
    out_path = tmp_path / "plan.csv"
    rc, out, err = _run(
        [
            "export",
            "--period",
            "2026-07",
            "--view",
            "plan",
            "--out",
            str(out_path),
            "--db",
            str(empty_db),
        ],
        capsys,
    )
    assert rc == 0, err
    text = out_path.read_text(encoding="utf-8")
    assert "Period,Category,Currency,Kind,Amount,Note" in text
    assert "Їжа/Ресторани" in text
    assert "-9000.0" in text


def test_export_variance_still_default(
    empty_db: Path, tmp_path: Path, capsys: pytest.CaptureFixture[str]
) -> None:
    """Default --view is still variance. Backward compat with v0.5.0
    callers."""
    _seed(empty_db, "2026-07", ("Їжа/Ресторани", 980, "baseline", -900000))
    out_path = tmp_path / "variance.csv"
    rc, out, err = _run(
        [
            "export",
            "--period",
            "2026-07",
            "--out",
            str(out_path),
            "--db",
            str(empty_db),
        ],
        capsys,
    )
    assert rc == 0, err
    assert out["view"] == "variance"
