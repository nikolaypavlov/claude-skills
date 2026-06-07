"""``pf-budget export`` (variance CSV/XLSX) + pf-report --budget block."""

from __future__ import annotations

import csv
import json
from datetime import UTC, datetime
from pathlib import Path

import pytest

from pf_skill.budget_cli import main as budget_main
from pf_skill.common import budget as bud
from pf_skill.common.reports import _try_period_string, build_report_bundle
from pf_skill.common.store import open_db
from pf_skill.report import main as report_main


def _run(argv: list[str], capsys: pytest.CaptureFixture[str]) -> tuple[int, dict, dict]:
    rc = budget_main(argv)
    captured = capsys.readouterr()
    out: dict = json.loads(captured.out) if captured.out.strip() else {}
    err: dict = {}
    if captured.err.strip():
        try:
            err = json.loads(captured.err.splitlines()[-1])
        except json.JSONDecodeError:
            err = {"raw": captured.err}
    return rc, out, err


def _seed_budget(db: Path, period: str = "2023-11") -> None:
    conn = open_db(db)
    bud.materialise_budget(
        conn,
        period=period,
        rows=[
            bud.PlanRow(period, "Підписки/Інше", 980, "baseline", -25000),
            bud.PlanRow(period, "Test/Other", 980, "baseline", -10000),
        ],
        source="seed",
    )
    conn.close()


# --- export CSV --------------------------------------------------------------


def test_export_csv_writes_variance_rows(
    empty_db: Path, tmp_path: Path, capsys: pytest.CaptureFixture[str]
) -> None:
    _seed_budget(empty_db, "2026-06")
    out_path = tmp_path / "variance.csv"
    rc, out, err = _run(
        [
            "export",
            "--period",
            "2026-06",
            "--out",
            str(out_path),
            "--db",
            str(empty_db),
        ],
        capsys,
    )
    assert rc == 0, err
    assert out["format"] == "csv"
    assert out_path.exists()
    with out_path.open(encoding="utf-8") as f:
        reader = csv.DictReader(f)
        rows = list(reader)
    assert len(rows) == 2
    headers = list(reader.fieldnames or [])
    assert headers == ["Period", "Category", "Currency", "Target", "Actual", "Delta", "% used"]
    # Target/Actual rendered as floats in major units (not minor).
    pi = next(r for r in rows if r["Category"] == "Підписки/Інше")
    assert float(pi["Target"]) == -250.0


def test_export_xlsx_writes_sheet(
    empty_db: Path, tmp_path: Path, capsys: pytest.CaptureFixture[str]
) -> None:
    pytest.importorskip("openpyxl")
    from openpyxl import load_workbook

    _seed_budget(empty_db, "2026-06")
    out_path = tmp_path / "variance.xlsx"
    rc, out, err = _run(
        [
            "export",
            "--period",
            "2026-06",
            "--out",
            str(out_path),
            "--db",
            str(empty_db),
        ],
        capsys,
    )
    assert rc == 0, err
    assert out["format"] == "xlsx"
    assert out_path.exists()
    wb = load_workbook(out_path)
    assert "Variance" in wb.sheetnames


def test_export_format_auto_from_extension(
    empty_db: Path, tmp_path: Path, capsys: pytest.CaptureFixture[str]
) -> None:
    _seed_budget(empty_db, "2026-06")
    out_path = tmp_path / "variance.xlsx"
    pytest.importorskip("openpyxl")
    rc, out, err = _run(
        [
            "export",
            "--period",
            "2026-06",
            "--out",
            str(out_path),
            "--format",
            "auto",
            "--db",
            str(empty_db),
        ],
        capsys,
    )
    assert rc == 0, err
    assert out["format"] == "xlsx"


# --- _try_period_string -------------------------------------------------------


def test_try_period_string_matches_exact_calendar_month() -> None:
    start = int(datetime(2026, 6, 1, tzinfo=UTC).timestamp())
    end = int(datetime(2026, 7, 1, tzinfo=UTC).timestamp())
    assert _try_period_string(start, end) == "2026-06"


def test_try_period_string_year_rollover() -> None:
    start = int(datetime(2026, 12, 1, tzinfo=UTC).timestamp())
    end = int(datetime(2027, 1, 1, tzinfo=UTC).timestamp())
    assert _try_period_string(start, end) == "2026-12"


def test_try_period_string_rejects_partial_month() -> None:
    start = int(datetime(2026, 6, 1, tzinfo=UTC).timestamp())
    end = int(datetime(2026, 6, 15, tzinfo=UTC).timestamp())
    assert _try_period_string(start, end) is None


def test_try_period_string_rejects_multi_month() -> None:
    start = int(datetime(2026, 6, 1, tzinfo=UTC).timestamp())
    end = int(datetime(2026, 8, 1, tzinfo=UTC).timestamp())
    assert _try_period_string(start, end) is None


# --- pf-report integration ----------------------------------------------------


def test_report_includes_budget_block_when_period_matches(
    mixed_currency_db: Path, capsys: pytest.CaptureFixture[str]
) -> None:
    """The fixture's mono_transactions ts is around 2023-11-14. Build
    a Nov 2023 budget, ask pf-report for that month, expect the
    budget block in the bundle."""
    conn = open_db(mixed_currency_db)
    conn.execute(
        "INSERT INTO tx_category (tx_id, category, set_at, set_by) "
        "VALUES ('uah_patreon', 'Підписки/Інше', 1700000050, 'manual')"
    )
    bud.materialise_budget(
        conn,
        period="2023-11",
        rows=[bud.PlanRow("2023-11", "Підписки/Інше", 980, "baseline", -25000)],
        source="seed",
    )
    conn.close()
    rc = report_main(
        [
            "--from",
            "2023-11-01",
            "--to",
            "2023-12-01",
            "--db",
            str(mixed_currency_db),
        ]
    )
    captured = capsys.readouterr()
    assert rc == 0, captured.err
    payload = json.loads(captured.out)
    assert "budget" in payload
    assert payload["budget"]["period"] == "2023-11"
    block_uah = next(b for b in payload["budget"]["blocks"] if b["currency_code"] == 980)
    cats = {line["category"]: line for line in block_uah["lines"]}
    assert cats["Підписки/Інше"]["actual_minor"] == -21232


def test_report_omits_budget_block_when_partial_month(
    mixed_currency_db: Path, capsys: pytest.CaptureFixture[str]
) -> None:
    conn = open_db(mixed_currency_db)
    bud.materialise_budget(
        conn,
        period="2023-11",
        rows=[bud.PlanRow("2023-11", "A", 980, "baseline", -100)],
        source="seed",
    )
    conn.close()
    report_main(
        [
            "--from",
            "2023-11-01",
            "--to",
            "2023-11-15",  # partial month
            "--db",
            str(mixed_currency_db),
        ]
    )
    captured = capsys.readouterr()
    payload = json.loads(captured.out)
    assert "budget" not in payload


def test_report_omits_budget_block_when_no_budget(
    mixed_currency_db: Path, capsys: pytest.CaptureFixture[str]
) -> None:
    report_main(
        [
            "--from",
            "2023-11-01",
            "--to",
            "2023-12-01",
            "--db",
            str(mixed_currency_db),
        ]
    )
    captured = capsys.readouterr()
    payload = json.loads(captured.out)
    assert "budget" not in payload


def test_build_report_bundle_directly(mixed_currency_db: Path) -> None:
    """Sanity check the bundle helper without going through CLI."""
    conn = open_db(mixed_currency_db)
    bud.materialise_budget(
        conn,
        period="2023-11",
        rows=[bud.PlanRow("2023-11", "A", 980, "baseline", -100)],
        source="seed",
    )
    bundle = build_report_bundle(
        conn,
        from_ts=int(datetime(2023, 11, 1, tzinfo=UTC).timestamp()),
        to_ts=int(datetime(2023, 12, 1, tzinfo=UTC).timestamp()),
    )
    assert "budget" in bundle
    assert bundle["budget"]["period"] == "2023-11"
    conn.close()
