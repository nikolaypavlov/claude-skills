"""Format-detection tests. Make sure the sniff is strict enough to reject
arbitrary XLSX files but tolerant of cosmetic title-row changes."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook
from privat24_import.parsers.detect import Format, detect


def test_detect_matches_generated_sample(sample_xlsx: Path) -> None:
    det = detect(sample_xlsx)
    assert det.fmt is Format.WEB_XLSX
    assert "privat24" in det.reason.lower() or "web xlsx" in det.reason.lower()


def test_detect_rejects_non_xlsx(tmp_path: Path) -> None:
    csv_path = tmp_path / "data.csv"
    csv_path.write_text("a,b,c\n1,2,3\n")
    det = detect(csv_path)
    assert det.fmt is Format.UNKNOWN
    assert "extension" in det.reason


def test_detect_rejects_xlsx_with_wrong_headers(tmp_path: Path) -> None:
    out = tmp_path / "other.xlsx"
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.append(["some title"])
    ws.append(["wrong", "header", "row"])
    ws.append(["body", 1, 2])
    wb.save(out)
    det = detect(out)
    assert det.fmt is Format.UNKNOWN


def test_detect_tolerant_to_title_text_change(
    tmp_path: Path, sample_xlsx: Path
) -> None:
    """The title row (row 1) is free-form Ukrainian; we should rely only
    on the header row (row 2). Mutate the title and re-detect."""
    wb = load_workbook(sample_xlsx)
    ws = wb.active
    assert ws is not None
    ws["A1"] = "Custom title text that differs"
    out = tmp_path / "title-changed.xlsx"
    wb.save(out)
    wb.close()
    det = detect(out)
    assert det.fmt is Format.WEB_XLSX
