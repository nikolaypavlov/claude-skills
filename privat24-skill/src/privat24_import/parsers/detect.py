"""Sniff a Privat24 statement export to decide which parser to use.

Today we recognise one format - ``web_xlsx``, the .xlsx download from
https://privat24.ua/statement. New formats (mobile app, FOP web cabinet)
will land here as additional ``Format`` variants without changing the
public detect API.
"""

from __future__ import annotations

from dataclasses import dataclass
from enum import Enum
from pathlib import Path


class Format(str, Enum):
    WEB_XLSX = "web_xlsx"
    UNKNOWN = "unknown"


@dataclass(frozen=True)
class Detection:
    fmt: Format
    reason: str


# Web-cabinet XLSX header. We match on the second row (zero-indexed 1)
# because the first row carries the human-readable "Історія операцій за
# період ..." title which would drift across exports.
WEB_XLSX_HEADERS: tuple[str, ...] = (
    "Дата",
    "Категорія",
    "Картка",
    "Опис операції",
    "Сума в валюті картки",
    "Валюта картки",
    "Сума в валюті транзакції",
    "Валюта транзакції",
    "Залишок на кінець періоду",
    "Валюта залишку",
)


def detect(path: str | Path) -> Detection:
    """Inspect ``path`` and return the best-guess ``Detection``."""
    path = Path(path)
    suffix = path.suffix.lower()
    if suffix not in {".xlsx", ".xlsm"}:
        return Detection(Format.UNKNOWN, f"unsupported extension '{suffix}'")
    # Import lazily so the module is importable in environments without
    # openpyxl (e.g. CI before deps are installed).
    from openpyxl import load_workbook  # type: ignore[import-not-found]

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        rows = ws.iter_rows(min_row=1, max_row=2, values_only=True)
        rows_list = list(rows)
    finally:
        wb.close()
    if len(rows_list) < 2:
        return Detection(Format.UNKNOWN, "too few rows")
    header = tuple((c or "").strip() for c in rows_list[1][: len(WEB_XLSX_HEADERS)])
    if header == WEB_XLSX_HEADERS:
        return Detection(Format.WEB_XLSX, "matched privat24 web XLSX header")
    return Detection(
        Format.UNKNOWN,
        f"row 2 did not match expected privat24 web XLSX header; got {header!r}",
    )
