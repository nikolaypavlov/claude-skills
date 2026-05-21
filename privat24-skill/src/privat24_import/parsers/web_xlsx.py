"""Parse the Privat24 web-cabinet XLSX export.

Shape (10 columns, see ``parsers.detect.WEB_XLSX_HEADERS`` for the header
text used to sniff this format):

    0  Дата                          datetime  DD.MM.YYYY HH:MM:SS, kyiv local
    1  Категорія                     str       Privat-assigned merchant category
    2  Картка                        str       masked PAN "5168 **** **** 3494"
    3  Опис операції                 str       free-text description
    4  Сума в валюті картки          Decimal   signed; negative = outflow
    5  Валюта картки                 str       3-letter ISO (UAH, USD, ...)
    6  Сума в валюті транзакції      Decimal   ABSOLUTE value, no sign
    7  Валюта транзакції             str       3-letter ISO; may differ from col 5 (FX)
    8  Залишок на кінець періоду     Decimal   balance after the txn
    9  Валюта залишку                str       balance currency (typically == col 5)

A single XLSX always covers one card (column 2 is constant). The
parser groups all rows into one account, deriving ``account_id`` from
the masked PAN so multiple exports of the same card merge correctly in
the store.
"""

from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal
from pathlib import Path
from typing import Iterator
from zoneinfo import ZoneInfo

from ..core.currencies import to_numeric

# Privat24 web exports use Europe/Kyiv local time (cells are naive). We
# attach ZoneInfo("Europe/Kyiv") at parse time so the resulting unix
# timestamp is true UTC seconds - downstream consumers can shift back to
# any timezone at query time.
KYIV_TZ = ZoneInfo("Europe/Kyiv")


@dataclass(frozen=True)
class ParsedRow:
    """One statement row after parsing. Stays close to the source so the
    caller can still build the natural-key hash."""

    ts: int  # unix seconds, true UTC
    category: str
    masked_pan: str
    description: str
    amount_minor: int  # signed; -50_00 = -50.00 UAH
    currency_code: int  # ISO 4217 numeric, account currency
    op_amount_minor: int | None  # signed; populated iff FX (ccy differs)
    op_currency_code: int | None
    balance_minor: int | None
    balance_currency_code: int | None
    raw: dict[str, object]  # source row for raw_json storage


@dataclass(frozen=True)
class ParsedStatement:
    """Aggregate of one parsed XLSX. All rows share ``account_id`` and
    ``account_currency_code`` (Privat24 exports one card per file)."""

    account_id: str
    masked_pan: str
    account_currency_code: int
    rows: list[ParsedRow]


_DATETIME_FMT = "%d.%m.%Y %H:%M:%S"
_PAN_SANITISER = re.compile(r"[^0-9*]")


def _to_unix_utc(value) -> int:
    """Convert a Privat24 date cell to a true UTC unix timestamp.

    Cells come through openpyxl as either a ``datetime`` (when the
    column is correctly typed) or a string in ``DD.MM.YYYY HH:MM:SS``
    form. Per the export convention they represent Europe/Kyiv local
    time, so we attach ``KYIV_TZ`` to the naive value and let
    ``datetime.timestamp()`` produce true UTC seconds (this also handles
    DST correctly via zoneinfo's historical tz data).

    If the cell is already tz-aware we trust the embedded offset.
    """
    if isinstance(value, datetime):
        dt = value
    elif isinstance(value, str):
        dt = datetime.strptime(value.strip(), _DATETIME_FMT)
    else:
        raise ValueError(
            f"unsupported date cell type {type(value).__name__}: {value!r}"
        )
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=KYIV_TZ)
    return int(dt.timestamp())


def _to_minor(value) -> int:
    """Round Decimal/float currency value to integer minor units.

    Privat24 statements show amounts to 2 decimals; we multiply by 100
    and round to int. Going via Decimal avoids the classic
    ``int(0.1 * 100)`` -> 9 problem with binary floats.
    """
    if value is None:
        raise ValueError("amount cell is empty")
    d = Decimal(str(value))
    return int((d * 100).quantize(Decimal("1")))


def _account_id_from_pan(masked_pan: str) -> str:
    """Stable account id derived from the masked PAN.

    The same card will produce the same id across imports. We strip
    spaces / asterisks separators so the id is filename-safe and stable
    against trivial formatting differences.
    """
    clean = _PAN_SANITISER.sub("", masked_pan or "")
    if not clean:
        raise ValueError(f"masked pan is empty or unparseable: {masked_pan!r}")
    return f"privat_pan_{clean}"


def parse(path: str | Path) -> ParsedStatement:
    """Parse the entire XLSX. Raises if the file is empty or contains
    multiple cards (we don't support multi-card exports yet)."""
    from openpyxl import load_workbook  # type: ignore[import-not-found]

    path = Path(path)
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        rows = list(_yield_rows(ws))
    finally:
        wb.close()
    if not rows:
        raise ValueError(f"{path}: no data rows after header")
    # Privat24 exports one card per file. We enforce that to avoid silent
    # account mixups; if this ever changes we'll need a multi-account
    # iteration here.
    cards = {r.masked_pan for r in rows}
    if len(cards) > 1:
        raise ValueError(
            f"{path}: expected exactly one card per export, got {sorted(cards)}"
        )
    masked_pan = rows[0].masked_pan
    account_currency = rows[0].currency_code
    if any(r.currency_code != account_currency for r in rows):
        raise ValueError(
            f"{path}: account currency drifts across rows; "
            "Privat24 statements should be per-account"
        )
    return ParsedStatement(
        account_id=_account_id_from_pan(masked_pan),
        masked_pan=masked_pan,
        account_currency_code=account_currency,
        rows=rows,
    )


def _yield_rows(ws) -> Iterator[ParsedRow]:
    """Iterate body rows, skipping the title row and header row.

    The XLSX has:
      row 1 = "Історія операцій за період ..."
      row 2 = column headers
      row 3+ = data
    """
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 2:
            continue
        if row is None or row[0] is None:
            continue
        (
            date_cell,
            category,
            masked_pan,
            description,
            amount,
            currency,
            op_amount,
            op_currency,
            balance,
            balance_currency,
        ) = row[:10]
        ts = _to_unix_utc(date_cell)
        amount_minor = _to_minor(amount)
        currency_code = to_numeric(currency or "")
        # Operation amount in the source is unsigned. Re-derive the sign
        # from the account amount so the cross-plugin convention
        # (negative = outflow) holds for both columns. A zero account
        # amount means we cannot derive a sign at all - store the op
        # amount as zero too so the pair stays internally consistent.
        op_currency_code = to_numeric(op_currency) if op_currency else None
        if op_currency_code is not None and op_currency_code != currency_code:
            op_amount_unsigned = abs(_to_minor(op_amount))
            if amount_minor == 0:
                op_amount_minor: int | None = 0
            else:
                sign = -1 if amount_minor < 0 else 1
                op_amount_minor = sign * op_amount_unsigned
        else:
            op_amount_minor = None
            op_currency_code = None
        balance_minor = _to_minor(balance) if balance is not None else None
        balance_currency_code = (
            to_numeric(balance_currency) if balance_currency else None
        )
        raw = {
            "date_raw": str(date_cell),
            "category": category,
            "masked_pan": masked_pan,
            "description": description,
            "amount": str(amount),
            "currency": currency,
            "op_amount": str(op_amount),
            "op_currency": op_currency,
            "balance": str(balance) if balance is not None else None,
            "balance_currency": balance_currency,
        }
        yield ParsedRow(
            ts=ts,
            category=category or "",
            masked_pan=masked_pan or "",
            description=description or "",
            amount_minor=amount_minor,
            currency_code=currency_code,
            op_amount_minor=op_amount_minor,
            op_currency_code=op_currency_code,
            balance_minor=balance_minor,
            balance_currency_code=balance_currency_code,
            raw=raw,
        )
